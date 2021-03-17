import colorama
from colorama import Fore, Back, Style
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Font,Alignment,PatternFill, colors
from tqdm.notebook import tqdm
import pandas as pd
import numpy as np
import datetime
import psycopg2
import pymysql
import time
import os
import re

from pymysql.err import IntegrityError, OperationalError

class EmptyException(BaseException):
    def __init__(self, message=None):
        pass
        

def revise_workbook(loc, cols, inds, sheets):
    """
    调整表格字体格式 & 单元格间距 & 边框 & 居中展示
    Params:
        loc: the location of the excel file
        cols: we actually care the number of columns
        inds: we actually care the number of data
        sheets: to revise the format among the given sheets
    """
    wb = load_workbook(loc)
    for sheet in sheets:
        ws = wb[sheet]
        #右下角最后一个单元格位置
        last_letter = chr(ord('A') + len(cols))
        last_row = str(1+len(inds))
        last_loc = last_letter + last_row

        # 边框
        thin = Side(border_style="thin", color="000000")#边框样式，颜色
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置

        #字体
        cell_font = Font(size=10, bold=False, name='Microsoft YaHei',  color="101013") # 数据
        init_cell_font = Font(size=10, bold=True, name='Microsoft YaHei', color='101013') # 首行首列
        #居中
        cell_align = Alignment(horizontal='center',vertical='center',wrap_text=True)

        #统一单元格格式
        for row in ws[f'A1:{last_loc}']:
            for cell in row:
                cell.border = cell_border
                cell.alignment = cell_align
                if cell.column_letter == 'A' or cell.row == 1:
                    cell.font = init_cell_font
                else:
                    cell.font = cell_font
        #设置首列宽度
        ws.column_dimensions['A'].width=24
        #设置其余列宽度
        for col in range(ord('A')+1, ord('A') + len(cols)+1):
            ws.column_dimensions[chr(col)].width=12
        #设置所有行高度
        for row in range(1,1+1+len(inds)):
            ws.row_dimensions[row].height=18
        wb.save(loc)
        
def diy(text, fore=None, back=None, style=None):
    """
    Desc:
        Enable the user to print a colorful text
    Params:
        fore: gimme the font color 
        back: set the background color for your text
        style: dim/bright/normal
    """
    if fore:
        fore = fore.lower()
    if back:
        back = back.lower()
    if style:
        style = style.lower()
    fore_dic = {'black': '\x1b[30m', 'red': '\x1b[31m', 
                'green': '\x1b[32m', 'yellow': '\x1b[33m', 
                'blue': '\x1b[34m', 'magenta': '\x1b[35m', 
                'cyan': '\x1b[36m', 'white': '\x1b[37m'}

    back_dic = {'black': '\x1b[40m', 'red': '\x1b[41m', 
                'green': '\x1b[42m', 'sellow': '\x1b[43m', 
                'blue': '\x1b[44m', 'magenta': '\x1b[45m', 
                'cyan': '\x1b[46m', 'white': '\x1b[47m'}

    style_dic = {'dim': '\x1b[2m', 'normal': '\x1b[22m', 
                 'bright': '\x1b[1m', 'reset_all': '\x1b[0m'}
    print(fore_dic.get(fore,'') + back_dic.get(back,'') + style_dic.get(style,'') + text + Style.RESET_ALL)
    
def cursor_into_df(my_cursor=None, close=False):
    """
    Desc:
        This function feteched data from cursor() and transformed into pandas.DataFrame()
    
    Params:
        my_cursor: pymysql.connect(**config).cursor()
        
    Returns:
        pandas.DataFrame()
    """
    columns = [i[0] for i in my_cursor.description]
    df = pd.DataFrame(my_cursor)
    try:
        df.columns = columns
    except:
        return pd.DataFrame()
    if close:
        my_cursor.close()
    return df

def get_first_day(week, year):
    """
    获取 year 第 week 周的第一天
    e.g.
        month = 12
        year = 2020
    
    """
    week, year = str(week), str(year)
    curr_year_first_day = year + "0101"
    yearstart = datetime.datetime.strptime(curr_year_first_day,'%Y%m%d')
    yearstartcalendarmsg = yearstart.isocalendar() 
    yearstartweek = yearstartcalendarmsg[1]  
    yearstartweekday = yearstartcalendarmsg[2]
    yearstartyear = yearstartcalendarmsg[0]
    if yearstartyear < int(year):
        daydelat = (8-int(yearstartweekday))+(int(week)-1)*7
    else :
        daydelat = (8-int(yearstartweekday))+(int(week)-2)*7
     
    a = (yearstart+datetime.timedelta(days=daydelat)).date()
    return a

def get_last_day(week, year):
    week = int(week)
    return get_first_day(week+1, year) - datetime.timedelta(days=4)

def clean_tt(x):
    """
    Desc:
        Remove specific suffix for one string,
            <who belongs to a column that looks like : FAT - Task_type - P - R>
            
            "[For RBS Proactive Only] " --> ""
            "(Proactive)" --> ""
            
    Params:
        x: string
    """
    
    x = x.replace('(Proactive)', '')
    if '|' in x:
        return x.split('|')[0]
    x = x.replace('[For RBS Proactive Only] ', '')
    x = x.strip()
    valid = re.compile(" PB\d{2}")
    if valid.search(x):
        last_ind = valid.search(x).span()[0]
        x = x[:last_ind]
    return x

def split_winston_p(x=None):
    if "For RBS Proactive Only" in x:
        return "Proactive-P"
    return "Proactive-R"

def enrich_df_tasks(df=None, tasks=None):
    """
    有些 CA 只负责一部分 Task，该函数负责填充数据保证其余 Task 也显示在表格中，且该 CA 对应该 Task 的值为0
    必须存在的列名：
        Task
        CA
    """
    to_be_appended = []
    if not tasks:
        tasks = list(df.Task.unique())
    for task in tasks:
        for CA in df.CA.unique():
            tmp = df.loc[(df.CA == CA) & (df.Task == task), :]
            if len(tmp) == 0:
                to_be_appended.append({"CA":CA,"Task":task})
    for d in to_be_appended:
        df = df.append(d, ignore_index=True)
    return df.sort_values(by=['CA', 'Task']).fillna(value=0)

def enrich_CM_tasks(df=None, tasks=None):
    """
    有些 CM 只负责一部分 Task，该函数负责填充数据保证其余 Task 也显示在表格中，且该 CM 对应该 Task 的值为0
    必须存在的列名：
        Task
        CM
    """
    to_be_appended = []
    if not tasks:
        tasks = list(df.Task.unique())
    for task in tasks:
        for cm in df.CM.unique():
            tmp = df.loc[(df.CM == cm) & (df.Task == task), :]
            if len(tmp) == 0:
                to_be_appended.append({"CM":cm,"Task":task})
    for d in to_be_appended:
        df = df.append(d, ignore_index=True)
    return df.sort_values(by=['CM', 'Task']).fillna(value=0)
    
def enrich_tasks(df=None, task_list=None):
    backup = []
    for task in task_list:
        if task not in list(df.Task.unique()):
            backup.append(task)
    for task in backup:
        df = df.append({'Task': task}, ignore_index=True)
    df = df.loc[df.Task != 'No Data']
    return df.sort_values(by=['Task']).fillna(value=0).reset_index(drop=True)

def enrich_overall_tasks(df=None, tasks=None):
    """
    有些 CM 只负责一部分 Task，该函数负责填充数据保证其余 Task 也显示在表格中，且该 CM 对应该 Task 的值为0
    必须存在的列名：
        Task
        CM
    """
    to_be_appended = []
    if not tasks:
        tasks = list(df.Task.unique())
        return df.sort_values(by=['Task']).fillna(value=0)
    for task in tasks:
        if task not in df.Task.unique():
            to_be_appended.append({"Task":task})
    for d in to_be_appended:
        df = df.append(d, ignore_index=True)
    return df.sort_values(by=['Task']).fillna(value=0)

def get_data(pro_cursor=None, twt_cursor=None, rs_cursor=None,
             q_wk_from=1, q_wk_to=2, q_year=None,
            the_CM=None, CM_login=None, 
            task_list_dir=None, not_wanted_status=None):
    """
    获取 特定 周数 期间 的 特定 CM 下的所有 CA 的数据
    需要制定 Task List
    
    Returns:
        touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt
    """
    print("0%",end="\r")
    # 1.1 Calculate Date data
    first_day_wk_from = get_first_day(q_wk_from, q_year) ## 计算得到 目标周数 的第一天和最后一天
    last_day_wk_to = get_last_day(q_wk_to, q_year)
    date_from = first_day_wk_from.strftime("%Y-%m-%d") # sql-日期起
    date_to = last_day_wk_to.strftime("%Y-%m-%d") # sql-日期止
    task_list = pd.read_excel(task_list_dir, sheet_name='Task FAT')
    winston_tasks = list(task_list.SOP.values)
    print("3%",end="\r")
    
    print("10% ...  Fetching Data",end="\r")
    # 2 Get Data
    # 2.1 执行 SQL 语句 -- Winston from RedShift
    rs_winston_query = f"""
    SELECT A.* FROM avs_ws.case_level_platform_addons A 
    WHERE A.region_id = 'JP-AVS'
    AND A.arrived_datetime BETWEEN '{date_from}' and '{date_to}'
    """
    rs_cursor.execute(rs_winston_query)
    winston_df = cursor_into_df(rs_cursor)
    print("15% ...  Fetching Data", end='\r')
    time.sleep(0.2)
    print("40% ...  Success", end='\r')
    # 第一次获取相关列的列名，不必再跑
    # col_names = []
    # for ind in [1,3,43,155,156,161,36,160]:
    #     col_names.append(winston_df.columns[ind])

    col_names = ['arrived_datetime',
     'min_resolved_datetime',
     'case_id_short',
     'asins_tobe_updated',
     'asins_updated',
     'task_type',
     'first_commenter',
     'last_resolved_by',
    ]
    wp_df = winston_df.loc[:, col_names]
    # 区分 proactive-R & proactive-P
    wp_df['PR'] = wp_df['task_type'].apply(lambda x:split_winston_p(x))
    wp_df['task_type'] = wp_df['task_type'].apply(lambda x: clean_tt(x))
    wp_wanted_cols = [
        'asins_tobe_updated', 'asins_updated',
        'task_type', 'last_resolved_by', 'PR'
    ]
    wp_df_use = wp_df.loc[:, wp_wanted_cols]
    wp_df_use = wp_df_use.rename(columns = {'task_type': 'Task', 'last_resolved_by':'CA'})
    wp_df_use['Task'] = wp_df_use['Task'].apply(lambda x: clean_tt(x))
    print("53%                 ",end='\r')
    ## 2.1.1 Filter out unnecessary Tasks
    wanted_rows_by_task = []
    for ind, row in wp_df_use.iterrows():
        if row.Task in winston_tasks:
            wanted_rows_by_task.append(ind)
    wp_df_use = wp_df_use.loc[wanted_rows_by_task, :].reset_index(drop=True)

    wp_df_asins_sum = wp_df_use.groupby(['CA', 'Task', 'PR']).sum().reset_index()
    wp_df_case_cnt = wp_df_use.groupby(['CA', 'Task']).count().PR.reset_index().rename(columns = {'PR': 'Cases_count'})
    print("60%",end="\r")
    # 2.2 SQL 代码 -- TWT from MySQL
    # CMs = ['tingyjin', 'xjinghui', 'wenjinzh']
    my_sql_1 = f"""SELECT * FROM TWT.TWT WHERE CM = '{CM_login}' """ 
    my_sql_2 = f""" AND DATE_FORMAT(Date,'%Y-%m-%d') BETWEEN '{date_from}' AND '{date_to}'"""
    my_sql = my_sql_1 + my_sql_2
    print("66% ...  Fetching TWT Data",end="\r")
    twt_cursor.execute(my_sql) # Execute SQL
    twt_df = cursor_into_df(twt_cursor) # 将查询结果转换成 pandas.DataFrame
    print("82% ...  Success", end='\r')
    # 2.2.1 Data Cleaning
    twt_df['Task'] = None # Initiate the column 'Task' in twt_df
    # 2.2.1.1 利用 Reg-Ex 找到 Task 名字以及之后的字符串内容，并放入新建的 Task 列
    target_ind = []
    valid = re.compile("JP\d{3}")
    for ind, row in twt_df.iterrows():
        s = row['Activity']
        if valid.search(s):
            str_ind = valid.search(s).span()[0]
            Task = s[str_ind:]
            twt_df.loc[ind, 'Task'] = Task
    twt_df = twt_df.dropna(subset = ['Task'])
    twt_df.reset_index(drop=True, inplace=True)
    twt_wanted_cols = [
        'Login', 'Start_time', 'End_time', 'Task'
    ]
    twt_df = twt_df.loc[:, twt_wanted_cols]

    # 2.2.1.2 计算 Hour
    twt_df['Hour'] = 0
    for ind, row in twt_df.iterrows():
        try:
            val = round((row['End_time'] - row['Start_time']).seconds/3600, 2)
            twt_df.loc[ind, 'Hour'] = val
        except:
            continue
    twt_df = twt_df.loc[:, ['Login', 'Task', 'Hour']]
    twt_df = twt_df.groupby(['Login', 'Task']).sum().reset_index()
    twt_df = twt_df.rename(columns={'Login':'CA'})
    print("90%                              ",end="\r")
    
    # 3. Total Data Cleaning
    # 3.1 Read Data from Touched  --> touched_df 
    sql = f"""
    SELECT *
    FROM Touched
    WHERE Allocation_week >= {q_wk_from} AND Allocation_week <= {q_wk_to}
        AND Allocation_year = {q_year}
    """
    pro_cursor.execute(sql)
    touched_tmp_df = cursor_into_df(pro_cursor)
    if len(touched_tmp_df) != 0:
        print("95%",end="\r")
        exec_py = """touched_tmp_df = touched_tmp_df.loc["""
        for i in not_wanted_status:
            exec_py += f"""(touched_tmp_df['Status'] != "{i}") & """
        exec_py = exec_py[:-3] + ", :]"
        exec(exec_py)
        tmp = touched_tmp_df.groupby(['Login', 'Task', 'Status']).ASIN.count().reset_index()
        touched_df = tmp.groupby(['Login', 'Task']).ASIN.sum().reset_index()
        touched_df = touched_df.rename(columns={'Login':'CA'})
    else:
        raise Exception(f"{q_wk_from} - {q_wk_to} 期间的数据为空，请上传数据!")
        touched_df = pd.DataFrame(columns=['CA', 'Task', 'ASIN'], data=[["No Data", "No Data", 0]])
    # 3.2 Read Data from Allocation and Management  --> allocation_df
    sql = f"""
        SELECT m.CM, a.*
        FROM Allocation a
            JOIN Management m
                ON a.CA = m.CA
        WHERE Week_from >= {q_wk_from}
            AND Week_to <= {q_wk_to}  
            AND Year = {q_year}
            AND m.CM = '{the_CM}'
    """

    allo_wanted_cols = [
        'CM', 'CA', 'Task', 'Basic_val', 'Range_val',
    ]

    pro_cursor.execute(sql)
    allocation_df = cursor_into_df(pro_cursor)
    if len(allocation_df) > 0:
        allocation_df = allocation_df.loc[:, allo_wanted_cols]
        # 3.3 Join allocation_df and touched_df  --> allo_touched
        allo_touched = pd.merge(allocation_df, touched_df,
                                left_on=['CA','Task'],
                                right_on=['CA','Task'], how='left').fillna(value=0)
        # 3.4 Join allo_touched and twt_df  --> atwt
        atwt = pd.merge(allo_touched, twt_df,
                       left_on = ['CA', 'Task'],
                       right_on = ['CA', 'Task'],   
                       how = 'left').fillna(value=0)
        # 3.5 Join atwt and wp_df_asins_sum  --> atwt_wp
        atwt_wp = pd.merge(atwt, wp_df_asins_sum, 
                          left_on = ['CA', 'Task'],
                          right_on = ['CA', 'Task'],
                          how = 'left').fillna(value=0)
        # 3.6 Join wp_df_case_cnt and tmp_atwt_wp(deduplicated from atwt_wp)  --> case_cnt
        tmp_atwt_wp = atwt_wp.drop_duplicates(['CA','Task']).loc[:, ['CM', 'CA', 'Task']]
        case_cnt = pd.merge(wp_df_case_cnt, tmp_atwt_wp,
                            left_on=['CA','Task'],
                            right_on = ['CA', 'Task'],
                            how = 'right').fillna(value=0)

        print(f"Data is 100% ready for {the_CM}({CM_login}) ranged from {q_year} Year {q_wk_from}th Week to {q_wk_to}th Week",end="\n")
        return touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt
    else:
        print(f"100% -- {the_CM}/{CM_login} -- Allocation Data(CM Data) is Empty in the MySQL database.",end="\n")
        return (touched_df, allocation_df,
                pd.DataFrame(columns=['CM', 'CA', 'Task', 'Basic_val', 'Range_val', 'ASIN'],
                             data = [["No Data","No Data","No Data",0,0,0]]), 
                pd.DataFrame(columns=['CM', 'CA', 'Task', 'Basic_val', 'Range_val', 'ASIN', 'Hour'],
                             data = [["No Data","No Data","No Data",0,0,0,0]]), 
                pd.DataFrame(columns=['CM', 'CA', 'Task', 'Basic_val', 'Range_val', 'ASIN', 'Hour', 'PR', 'asins_tobe_updated', 'asins_updated'], 
                             data = [["No Data","No Data","No Data", 0,0,0,0,"Proactive-R",0,0], 
                                     ["No Data","No Data","No Data", 0,0,0,0,"Proactive-R",0,0]]), 
                pd.DataFrame(columns=['CM', 'CA', 'Task', 'Cases_count'],
                             data = [["No Data","No Data","No Data", 0]]))

def get_all_data(pro_cursor=None, twt_cursor=None, rs_cursor=None,
                 q_wk_from=1, q_wk_to=2, q_year=None,
                 names=None, logins=None, 
                 task_list_dir=None, not_wanted_status=None):
    """
    Params:
        pro_cursor
        twt_cursor
        rs_cursor
        q_wk_from
        q_wk_to
        q_year
        names: List[CM 的名字]
        logins: List[CM 的 logins]
        task_list_dir
        
    遍历所有 CM 在目标时间周期内的所有数据，并返回几个有用的 List[List] 结构
    
    e.g.:
        the_CMs = ['Tingya', 'Jinghui', 'Wenjing']
        CMs_login = ['tingyjin', 'xjinghui', 'wenjinzh']
    """
    
    touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt = pd.DataFrame(),pd.DataFrame(),pd.DataFrame(),pd.DataFrame(),pd.DataFrame(),pd.DataFrame()
    # 1. 获取 该时间段的所有数据，并做相应清理
    for ind in range(len(logins)):
        the_CM, CM_login = names[ind], logins[ind]
        c_touched_df, c_allocation_df, c_allo_touched, c_atwt, c_atwt_wp, c_case_cnt = get_data(pro_cursor, twt_cursor, rs_cursor,q_wk_from, q_wk_to, q_year,the_CM, CM_login, task_list_dir, not_wanted_status)
        
        touched_df = pd.concat([touched_df, c_touched_df]).reset_index(drop=True)
        allocation_df = pd.concat([allocation_df, c_allocation_df]).reset_index(drop=True)
        allo_touched = pd.concat([allo_touched, c_allo_touched]).reset_index(drop=True)
        atwt = pd.concat([atwt, c_atwt]).reset_index(drop=True)
        atwt_wp = pd.concat([atwt_wp, c_atwt_wp]).reset_index(drop=True)
        case_cnt = pd.concat([case_cnt, c_case_cnt]).reset_index(drop=True)
    
#     return [
#         touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt
#     ]
    
    allocation_df_CM = allocation_df.groupby(['Task']).sum().reset_index()
    allo_touched_CM = allo_touched.groupby(['Task']).sum().reset_index()
    atwt_CM = atwt.groupby(['Task']).sum().reset_index()
    atwt_wp_CM = atwt_wp.groupby(['Task']).sum().reset_index()
    case_cnt_CM = case_cnt.groupby(['Task']).sum().reset_index()
    
    # 2. 进一步做数据清理，生成多个 Lists 
    # 2.1 Basic Goal
    pb_bg = allocation_df_CM
    pb_tasks = list(pb_bg.Task.unique())
    pb_tasks.sort()
    basic_goal_lists = [pb_tasks, ]
    for ind, task in enumerate(pb_tasks):
        basic_goal_lists.append([])
        tmp = list(pb_bg.loc[pb_bg.Task == task,:].Basic_val.values)
        tmp = [round(i,0) for i in tmp]
        basic_goal_lists[-1] = tmp

    # 2.2 Range Goal
    pb_rg = allocation_df_CM
    range_goal_lists = [pb_tasks,]
    for ind, task in enumerate(pb_tasks):
        range_goal_lists.append([])
        tmp = list(pb_rg.loc[pb_rg.Task == task,:].Range_val.values)
        tmp = [round(i,0) for i in tmp]
        range_goal_lists[-1] = tmp
        
    # 2.3 P-Completed ASINs
    CA_names = list(atwt_wp.CA.unique())
    p_completed_asin = atwt_wp.loc[atwt_wp.PR == 'Proactive-P'] \
                    .groupby(['Task']).sum().reset_index()
    p_completed_asin = enrich_tasks(p_completed_asin, pb_tasks)
    pca_lists = [pb_tasks,]
    for ind, task in enumerate(pb_tasks):
        pca_lists.append([])
        tmp = list(p_completed_asin.loc[p_completed_asin.Task == task,:].asins_updated.values)
        tmp = [round(i,4) for i in tmp]
        pca_lists[-1] = tmp

    # 2.4 P-Touched ASINs
    p_touched_new_df = allo_touched_CM
    pta_lists = [pb_tasks, ]
    for ind, task in enumerate(pb_tasks):
        pta_lists.append([])
        tmp = list(p_touched_new_df.loc[p_touched_new_df.Task == task,:].ASIN.values)
        tmp = [round(i,4) for i in tmp]
        pta_lists[-1] = tmp

    # 2.5 percent of basic goal
    pct_bg_lists =[pb_tasks, ]
    for ind, task in enumerate(pb_tasks):
        left, right = pca_lists[ind+1], basic_goal_lists[ind+1]
        curr_pct_bg = []
        for j in range(len(left)):
            if right[j] != 0:
                val = round(left[j]/right[j], 4)
            else:
                val = 0
            curr_pct_bg.append(val)
        pct_bg_lists.append(curr_pct_bg)
        
    # 2.6 percent of touched
    pct_tg_lists = [pb_tasks, ]
    for ind, task in enumerate(pb_tasks):
        left, right = pta_lists[ind+1], basic_goal_lists[ind+1]
        curr_pct_tg = []
        for j in range(len(left)):
            if right[j] != 0:
                val = round(left[j]/right[j], 4)
            else:
                val = 0
            curr_pct_tg.append(val)
        pct_tg_lists.append(curr_pct_tg)
        
    # 2.7 Reactive ASINs to be updated
    ra_to_be_df = atwt_wp.loc[atwt_wp.PR == 'Proactive-R'].groupby(['Task']).sum().reset_index()
    ra_to_be_df = enrich_tasks(ra_to_be_df, pb_tasks)
    ra_tobe_lists   = [pb_tasks, ]
    for ind, task in enumerate(pb_tasks):
        curr_ra_tobe = list(ra_to_be_df.loc[ra_to_be_df.Task == task,:].asins_tobe_updated.values)
        ra_tobe_lists.append(curr_ra_tobe)

    # 2.8  Reactive ASINs updated
    ra_ed_df = atwt_wp.loc[atwt_wp.PR == 'Proactive-R'].groupby(['Task']).sum().reset_index()
    ra_ed_df = enrich_tasks(ra_ed_df, pb_tasks)
    ra_ed_lists = [pb_tasks, ]
    for ind, task in enumerate(pb_tasks):
        curr_ra_ed = list(ra_ed_df.loc[ra_ed_df.Task == task,:].asins_updated.values)
        ra_ed_lists.append(curr_ra_ed)

    # 2.9 Completed ASINs (P + R)
    pr_ed_lists = [pb_tasks,]
    for ind, task in enumerate(pb_tasks):
        l, r = pca_lists[ind+1], ra_ed_lists[ind+1]
        curr_pr_ed = [round(l[j] + r[j], 2) if r[j] or l[j] else 0 for j in range(len(r))]
        pr_ed_lists.append(curr_pr_ed)

    # 2.10 Completion
    completion_lists = [pb_tasks,]
    for ind, task in enumerate(pb_tasks):
        l, r = pr_ed_lists[ind+1], basic_goal_lists[ind+1]
        curr_completion =[round(l[j]/r[j], 2) if r[j] !=0 else 0 for j in range(len(r))]
        completion_lists.append(curr_completion)
        
    return (touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt, 
                basic_goal_lists, range_goal_lists, pca_lists, pta_lists, 
                pct_bg_lists, pct_tg_lists, 
                ra_tobe_lists,ra_ed_lists, pr_ed_lists, 
                completion_lists
           )

def get_by_task_data(avs_pro_cursor=None, twt_cursor=None, rs_cursor=None,
                    weeks=None, years=None, the_CMs=None, CMs_login=None, task_list_dir=None, not_wanted_status=None):
    
    summarize_weeks_all = {}  
    overall_all = {}
    x_axis_titles = ['Overall'] + the_CMs
#     touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt = \
#         pd.DataFrame(),pd.DataFrame(),pd.DataFrame(),pd.DataFrame(),pd.DataFrame(),pd.DataFrame()
    group_cnt = len(weeks)
    
    for i in range(group_cnt):
        ans = \
            get_all_data(avs_pro_cursor, twt_cursor, rs_cursor, 
                         weeks[i][0], weeks[i][1], years[i], 
                         the_CMs, CMs_login, task_list_dir, not_wanted_status)
        origin_data_lists, by_task_data_lists = ans[:6], ans[6:]
#         basic_goal_lists, range_goal_lists \
#             ,pca_lists, pta_lists, pct_bg_lists \
#             , pct_tg_lists, ra_tobe_lists, ra_ed_lists \
#             , pr_ed_lists, completion_lists = all_data_lists

        for metric_ind, data in enumerate(by_task_data_lists):
            tasks, vals = data[0], data[1:]
            vals = [i[0] for i in vals]
            for task_ind, val in enumerate(vals):
                curr_task = tasks[task_ind]
                while curr_task not in summarize_weeks_all:
                    summarize_weeks_all[curr_task] = [
                        [0 for ii in range(group_cnt)] for jj in range(10)
                    ] # 单一 Task == 10个 metric，每行 有 group_cnt 列数据，初始为 0
                summarize_weeks_all[curr_task][metric_ind][i] = round(val, 2)
                
        overall_data_lists = get_overall_all_data(origin_data_lists) ## 10 个 Metrics
        for metric_ind, data in enumerate(overall_data_lists):
            
            for x_ind, x_val in enumerate(data):
                curr_x = x_axis_titles[x_ind] # 'overall' 或者 其他 CM
                if curr_x not in overall_all:
                    overall_all[curr_x] = [
                        [0 for ii in range(group_cnt)] for jj in range(10)
                    ] 
                overall_all[curr_x][metric_ind][i] = round(x_val, 2)
    return overall_all, summarize_weeks_all

def get_overall_all_data(overall_data_lists=None):
    touched_df, allocation_df_CM, allo_touched_CM, atwt_CM, atwt_wp_CM, case_cnt_CM = \
        overall_data_lists
    # Remove "No Data"
    allo_touched_CM = allo_touched_CM.loc[allo_touched_CM.CM != 'No Data']
    atwt_CM = atwt_CM.loc[atwt_CM.CM != 'No Data']
    atwt_wp_CM = atwt_wp_CM.loc[atwt_wp_CM.CM != 'No Data']
    case_cnt_CM = case_cnt_CM.loc[case_cnt_CM.CM != 'No Data']

    # Overall
    # 2.1 Basic Goal
    oa_bg = sum(allocation_df_CM.Basic_val.values)

    # 2.2 Range Goal
    oa_rg = sum(allocation_df_CM.Range_val.values)

    # 2.3 Proactive Completed ASIN count
    p_completed_asin = atwt_wp_CM.loc[atwt_wp_CM.PR == 'Proactive-P']
    oa_pca = sum(p_completed_asin.asins_updated.values)

    # 2.4 P-Touched ASINs
    oa_pta = sum(allo_touched_CM.ASIN.values)

    # 2.5 percent of basic goal
    oa_pct_bg = round(oa_pca / oa_bg, 2)

    # 2.6 percent of touched
    oa_pct_tg = round(oa_pta / oa_bg, 2)

    # 2.7 Reactive ASINs to be updated
    ra_to_be_df = atwt_wp_CM.loc[atwt_wp_CM.PR == 'Proactive-R']
    oa_ra_tobe = sum(ra_to_be_df.asins_tobe_updated.values)

    # 2.8  Reactive ASINs updated
    ra_ed_df = atwt_wp_CM.loc[atwt_wp_CM.PR == 'Proactive-R']
    oa_ra_ed = sum(ra_ed_df.asins_updated.values)

    # 2.9 Completed ASINs (P + R)
    oa_pr_ed = oa_pca + oa_ra_ed

    # 2.10 Completion
    oa_completion = round(oa_pr_ed / oa_bg, 2)
     
        
    # By CM & plus Overall
    # 2.1 Basic Goal
    cm_bg = list(allocation_df_CM.groupby('CM').sum().reset_index().sort_values(by='CM').Basic_val)
    cm_bg.insert(0, round(oa_bg, 2))
    # 2.2 Range Goal
    cm_rg = list(allocation_df_CM.groupby('CM').sum().reset_index().sort_values(by='CM').Range_val)
    cm_rg.insert(0, round(oa_rg, 2))
    # 2.3 Proactive Completed ASIN count
    cm_pca = list(atwt_wp_CM.loc[atwt_wp_CM.PR == 'Proactive-P'] \
        .groupby('CM').sum().reset_index() \
        .sort_values(by='CM') \
        .asins_updated)
    cm_pca.insert(0, round(oa_pca, 2))
    # 2.4 P-Touched ASINs
    cm_pta = list(allo_touched_CM.groupby('CM').sum().reset_index().sort_values(by='CM').ASIN)
    cm_pta.insert(0, oa_pta)
    # 2.5 percent of basic goal
    cm_pct_bg = [round(cm_pca[i]/cm_bg[i], 2) for i in range(len(cm_bg))]
    # 2.6 percent of touched
    cm_pct_tg = [round(cm_pta[i]/cm_bg[i], 2) for i in range(len(cm_bg))]
    # 2.7 Reactive ASINs to be updated
    cm_ra_tobe = list(atwt_wp_CM.loc[atwt_wp_CM.PR == 'Proactive-R'] \
                       .groupby('CM').sum().reset_index() \
                       .sort_values(by='CM') \
                       .asins_tobe_updated)
    cm_ra_tobe.insert(0, oa_ra_tobe)
    # 2.8  Reactive ASINs updated
    cm_ra_ed = list(atwt_wp_CM.loc[atwt_wp_CM.PR == 'Proactive-R'] \
                       .groupby('CM').sum().reset_index() \
                       .sort_values(by='CM') \
                       .asins_updated)
    cm_ra_ed.insert(0, oa_ra_ed)
    # 2.9 Completed ASINs (P + R)
    cm_pr_ed = [cm_pca[i] + cm_ra_ed[i] for i in range(len(cm_pca))]
    # 2.10 Completion
    cm_completion = [round(cm_pr_ed[i]/cm_bg[i], 2) for i in range(len(cm_bg))]
        
        
    oabt_data_lists = [cm_bg, cm_rg, cm_pca, cm_pta, cm_pct_bg, 
                       cm_pct_tg, cm_ra_tobe, cm_ra_ed, cm_pr_ed, cm_completion]
        
    return oabt_data_lists

def get_params(Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, total, ignore_params=0):
    while total%row_sheet_counts != 0:
        total+=1
    col_sheet_counts = total//row_sheet_counts ## 放多少列

    grid_height = (100 - padding_vert*2 - (row_sheet_counts-1)*grid_vert)/row_sheet_counts
    grid_width = (100 - padding_hori*2 - (col_sheet_counts-1)*grid_hori)/col_sheet_counts
    
    ignore_patch = [0 for i in range(ignore_params)]
    top_loc = []
    top_patch = [padding_vert + grid_height*i + grid_vert*i for i in range(row_sheet_counts)]
    for i in range(col_sheet_counts):
        top_loc.extend(top_patch)
    top_loc.sort()
    bottom_loc = [i for i in top_loc[::-1]]
    
    left_loc = []
    left_patch = [padding_hori + grid_width*i + grid_hori*i for i in range(col_sheet_counts)]
    for i in range(row_sheet_counts):
        left_loc.extend(left_patch)
    right_loc = [i for i in left_loc[::-1]]

    top_loc = ignore_patch + top_loc
    bottom_loc = ignore_patch + bottom_loc
    left_loc = ignore_patch + left_loc
    right_loc = ignore_patch + right_loc               
    # Grid 尺寸
    Grid_height = Grid_width*(row_sheet_counts/(1-0.02*padding_vert))/(col_sheet_counts/(1-0.02*padding_hori))
    
    return col_sheet_counts, top_loc, bottom_loc, left_loc, right_loc, Grid_height


