# dafault modules
import os
import re
import sys
import time
import random
import shutil
import camelot
import datetime
import colorama
import warnings
import threading
import numpy as np
import pandas as pd
warnings.filterwarnings('ignore')
from colorama import Fore, Back, Style

# notebook progress bar
from tqdm.notebook import tqdm

# Excel operations
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Font,Alignment,PatternFill, colors

# sql database
import psycopg2
import pymysql
from pymysql.err import IntegrityError, OperationalError

# pyecharts
from pyecharts import options as opts
from pyecharts.charts import Bar
from pyecharts.faker import Faker
from pyecharts.globals import ThemeType
from pyecharts.charts import Bar, Grid, Line, Tab
from snapshot_selenium import snapshot

# personal tool-kit
from proactive_KIT import (
    revise_workbook,
    diy, cursor_into_df, get_first_day, get_last_day, clean_tt, split_winston_p,
    enrich_df_tasks, enrich_CM_tasks, enrich_overall_tasks, enrich_tasks,
    get_data, get_all_data, get_by_task_data, get_overall_all_data, get_params
)

# pyqt5 GUI
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLineEdit, QLabel, QHBoxLayout, QVBoxLayout, qApp, \
    QDesktopWidget, QFileDialog, QPlainTextEdit, QTextEdit, QProgressBar, QTabWidget, QCheckBox
from PyQt5.QtCore import QCoreApplication, QTimer, Qt
from PyQt5.QtGui import QFont, QTextCursor, QIcon
from PyQt5.Qt import QThread, QMutex, pyqtSignal, QObject


#  EmailSender modules
import win32com.client as win32
import codecs
# emoji & winsound
import emoji
import winsound
my_emoji = emoji.EMOJI_UNICODE
### 'en', 'es', 'pt', 'it'


# QPushButton
#         border-color: #48A43F;
#         border-width: 1px;
#         border-radius: 10px;
#         padding: 1px;
run_button_Style = """
        QPushButton{
        text-align: center;
        color: white;
        background-color: #008754;
        font: bold;
        border-radius: 10px;
        height: 35px;
        border-style: outset;
        font: 20px;
        }
        QPushButton:hover{
        background-color: #48A43F;
        }
        QPushButton:pressed{
        text-align:center;
        background-color: #026A52;
        color:white;
        font: bold;
        border-radius: 10px;
        padding: 1px;
        height: 30px;
        border-style: outset;
        }
        QPushButton:checked{
        background-color: #AB2524;
        color:white;
        }
        """
white_button_Style =  """
        QPushButton{
        text-align: center;
        color: #000000;
        background-color: #FFFFFF;
        font: bold;
        height: 35px;
        border-style: outset;
        font: 18px;
        }
        QPushButton:hover{
        border-color: #48A43F;
        border-radius: 10px;
        border-width: 1px;
        }
        QPushButton:pressed{
        border-color:  #0A181F;
        border-width: 1px;
        border-radius: 10px;
        padding: 1px;
        height: 30px;
        border-style: outset;
        }
        """
exit_button_Style = """
    QPushButton{
    text-align: center;
    background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                      stop: 0 #f6f7fa, stop: 1 #dadbde);
    color: black;
    font:bold;
    border-color: grey;
    border-width: 2px;
    border-radius: 10px;
    padding: 6px;
    height: 30px;
    border-style: outset;
    font: 18px;
    }
    QPushButton:hover{
    background-color:#7E8B92;
    }
    QPushButton:pressed{
    background-color: black;
    border-style: inset;
    color: white;
    border-color: grey;
    border-width: 2px;
    border-radius: 10px;
    padding: 6px;
    border-style: inset;
    
    }
    """
open_folder_button_Style = """
    QPushButton{
    text-align: center;
    background-color: #008754;
    color: white;
    border-radius: 10px;
    padding: 6px;
    height: 20px;
    border-style: outset;
    }
    QPushButton:hover{
    background-color: #48A43F;
    }
    QPushButton:pressed{
    background-color: #026A52;
    border-style: inset;
    }
    """
info_tool_tip = """
    使用说明
    1. 报告导出路径
        储存 HTMLs & Data Frames
    2. Task List Excel File
        含有名为 Task FAT 的标签页的表格，其中 SOP 列包含的 Task 用于参考。
    3. CM Info Excel File
        含有现役 CMs 的名字缩写（格式请参考下一段 the_CM 的详细描述）以及 Login.
    4. Params Excel File
        旨在提供细微参数修改，比如图表展示的大小，预设的文件地址等
    5. By Person & Time Compare Report
        单周期报告，只包含一个 CM 及其 CAs 的数据
    6. Overall & By Task Report
        多周期报告，包含所有 CM 及其 CAs 的数据
"""
info_html = """
    <html>
    <body>
    <h1>
    使用说明
    </h1>
    <p>
    ⚪ 报告导出路径
        <pre>\t储存 HTMLs & Data Frames
        </pre>
    </p>
    <p>
    ⚪ Task List Excel File
        <pre>\t含有名为 Task FAT 的标签页的表格，其中 SOP 列包含的 Task 用于参考。
        </pre>
    </p>
    <p>
    ⚪ CM Info Excel File
        <pre>\t含有现役 CMs 的名字缩写（格式请参考下一段 the_CM 的详细描述）以及 Login。
        </pre>
    </p>
    <p>
    ⚪ Params Excel File
        <pre>\t旨在提供细微参数修改，比如图表展示的大小，预设的文件地址等
        </pre>
    </p>
    <p>
    ⚪ By Person & Time Compare Report
        <pre>\tthe_CM
            <pre>\t\t与 //CM Data 中的名字一致 e.g. Tingya</pre>
        </pre>
        <pre>\tCM_login
            <pre>\t\te.g. tingyjin</pre>
        </pre>
    </p>
    <p>
    ⚪ Overall & By Task Report
        <pre>\t为多组数据提供对比，需要输入不同的时间周期，其中不同CM的信息由CM Info Excel File 负责收取</pre>
    </p>
    </body>
    </html>       """#.replace("<pre>", "<p>").replace("</pre>", "</p>")

# QLabel
label_font = QFont()   #实例化字体对象
label_font.setFamily('Microsoft YaHei')  #字体
label_font.setBold(True)  #加粗
# font.setItalic(True)    #斜体
# font.setStrikeOut(True)  #删除线
# font.setUnderline(True)   #下划线
label_font.setPointSize(10)   #字体大小
#font.setWeight(25)   #可能是字体的粗细
# self.label.setFont(label_font)

# Colors
fill_1 = PatternFill("solid", fgColor="D4652F")
## 可视化色谱
colors = [
    '#E79C00',
    '#2B2C7C',
    '#F7BA0B',
    '#D9C022',
    '#8A5A83',
    '#7E292C',
    '#13447C',
    '#41678D',
    '#CB8D73',
    '#2B2C7C',
    '#5A3A29',
    '#887142',
    '#44322D',
    '#CB8D73',
    '#A02128',
]
## 可视化进度色谱字典
color_dict = {
    'warning':   "#A1232B", # 红色
    'success':   "#28713E", # 绿色
    'progress':  "#7E8B92", # 灰绿色
    'fetching':  "#8A5A83", # 紫色
    'data_ready':"#DD7907"  # 橙色
}
# #F3E03B # 黄色
# #FFA421 #亮橙色


#　获取当日时间，用作文件夹命名
moment = time.gmtime()
m = str(moment.tm_mon)
if len(m) == 1:
    m = "0" + m
d = str(moment.tm_mday)
if len(d) == 1:
    d = "0" + d
date_in_file_name = m+d
moment_year = time.gmtime().tm_year
print(f"Today: {moment_year}{date_in_file_name}")

# 连接 MySQL 数据库并写入数据  
def get_mysql_connection():
    mysql_connect = pymysql.connect(host='dev-dsk-wangting-2a-25bc2431.us-west-2.amazon.com', 
                              user='avs_user', 
                              password='avs_pro', 
                              database='AVS_proactive',
                              charset='utf8') #服务器名,账户,密码,数据库名
    mysql_cursor = mysql_connect.cursor()
    return mysql_connect, mysql_cursor

try:
    mysql_connect, mysql_cursor = get_mysql_connection()

    # 连接 RedShift 数据库
    rs_connect = psycopg2.connect(database="rsbidw",
                            user="avs_user",
                            password="A21sP22oR$",
                            host="rsbi-analytics.clszsz7jap6y.us-east-1.redshift.amazonaws.com",
                            port="8192")

    rs_cursor = rs_connect.cursor()
except Exception as e:
    print(e)
    print('Network Connection Failed')
    sys.exit()
    
    
# User's login
def get_curr_login():
    """
    含有 Downloads 且可以访问的文件夹名字即为你的 Login
    """
    path = r'C:\Users'
    for folder in os.listdir(path)[::-1]:
        nxt = os.path.join(path, folder)
        if os.path.isfile(nxt):
            continue
        try:
            for file in os.listdir(nxt)[::-1]:
                if file == 'Downloads':
                    return folder
        except:
            continue
    return None

# my_login = get_curr_login()
my_login = os.path.expanduser("~").split('\\')[-1]
    

qmut_init = QMutex()
qmut_report = QMutex()
qmut_single = QMutex()
qmut_multiple = QMutex()
qmut_integrate = QMutex()


class Thread_Email(QThread):
    _signal = pyqtSignal(str)
    def __init__(self, Inputs):
        super().__init__()
        self.Inputs = Inputs
    
    def run(self):
        # 取出 参数
        receivers, CCs, reports, subject, content = self.Inputs
        
        # 建立邮箱对象
        outlook = win32.Dispatch('Outlook.Application')
        mail_item = outlook.CreateItem(0)
        
        # 添加收件人
        for receiver in receivers:
            mail_item.Recipients.Add(receiver)
        # Copy 方
        mail_item.CC = CCs
        mail_item.Subject = subject
        mail_item.BodyFormat = 2  
            
        ## 邮件收尾： 发件人-部门-公司logo
        self.email_ending = \
            f"<br><p class='p4'>Regards<br>{my_login}<br>Amazon RBS AVS Team<br></p>"
        
        ## 邮件正文
        my_html  = \
            """<style type="text/css">
            html,
            body {
                text-align: left;
                margin:0;
                padding:0;
                background: #F7FBF5;
            }
            p {
                font: Microsoft YaHei
                font-size: 10px;
            }
            p span {
                font-weight: bold;
            }
            p div {
                font-size: 10px;
            }
            </style>
            """
        for ind, c in enumerate(content):
            my_html += f"<p class='p{ind+1}'>{c}</p>"
        
        ### 将该 HTML 文件保存并作为附件放在邮件中
        email_html = my_html + self.email_ending
        for report in reports:
            mail_item.Attachments.Add(report)
        
        mail_item.HTMLBody = email_html
        mail_item.Display()
        
        
class Thread_Scout(QThread):
    _signal = pyqtSignal(str)
    def __init__(self, Inputs):
        super().__init__()
        self.Inputs = Inputs
    
    def run(self):
        Task, Year = self.Inputs
        scout_query = f"""
            SELECT DISTINCT Allocation_week FROM Touched
            WHERE Task LIKE '{Task}%' AND Allocation_year = {Year}
        """
        mysql_cursor.execute(scout_query)
        scout_df = cursor_into_df(mysql_cursor)
        res = sorted(scout_df.Allocation_week.unique())
        res = [str(i) for i in res]
        output = f"{Task} 以下工作周已上传数据:\n\t" + ",".join(res)
        self._signal.emit(output)

    
class Thread_Upload_Touched(QThread):
    _signal = pyqtSignal(str)
    _bug_signal = pyqtSignal(str) #self._bug_signal.emit()
    _p_signal = pyqtSignal(int) # 进度信号 self._p_signal.emit()
    def __init__(self, Inputs):
        super().__init__()
        self.Inputs = Inputs

    def run(self):
        for ind, vals in enumerate(self.Inputs):
            a,b,c,d = vals
            self.process_read_upload(a,b,c,d)
        
    def read_task_dir(self, root, cols, task, wanted_week, wanted_file, in_files, in_weeks, files_in_mysql):
        """
        遍历该 task 文件夹下的所有文件，若满足以下条件则读取特定列数据:
            1. _2020.xlsx

        Params:
            root: 包含所有 Task 文件夹的文件夹
            cols: 提取所需的列
            task: 当前 Task
            wanted_week: List[Integers]
            wanted_file: List[String]
            in_files: 打开的 text 文件，用于记录已录入的数据文件名
            in_weeks: 字典 -- 对应的文件已读取的 week 数
            files_in_mysql: open("files_in_mysql.txt", 'a')
        """
        task_dir = root + "\\" + task
        valid_excel = re.compile("[0-9]\d{3}\.[xls]+$")
        all_df = pd.DataFrame()
        file_locs_tobe_uploaded = []

        wanted_all = False     
        if wanted_week == 'A':
            # If the user wants data of all weeks, we will filter out those in the record and upload the rest!
            wanted_all = True

        for file in os.listdir(task_dir):
            file_loc = task_dir + '\\' + file
            # Find files that meet requirements
            right_file = False

            if not valid_excel.findall(file) or "~$" in file:
                continue

            # According to param wanted_week, filter out data

            if file_loc in in_files:
                # 若 file_in_mysql 记录过该文件，则过滤掉记录中的 week
                wanted_left = [i for i in wanted_week if i not in in_weeks[file_loc]]
                if len(wanted_left) == 0:
                    continue
                wanted_week = wanted_left

            # According to the specific file name rule provided by users, filter out data
            for file_part in wanted_file:
                # Because we have multiple name parts for users to input, we need a for loop
                if file_part in file:
                    right_file = True
                    break

            if not right_file:
                continue

            file_locs_tobe_uploaded.append(file_loc)
            if file_loc not in in_weeks:
                in_weeks[file_loc] = []

        length_files = len(file_locs_tobe_uploaded)
        if length_files == 0:
            self._bug_signal.emit(f'{task} -- No data is ready to upload, please CHECK!')
            pos_reasons = "Three possible reasons:" + \
                "\t1. The files_in_mysql implies the data is uploaded before." + \
                "\t2. The rule for the name of file limits the options." + \
                "\t3. The name of file does not meet the No.1 rule(...{Year}....)"
            self._bug_signal.emit(pos_reasons)
            return False
        elif length_files > 1:
            self._signal.emit(f"\t{length_files} files will be uploaded:")
        else:
            self._signal.emit("\tOnly 1 file will be uploaded:")
        for ind, file_loc in enumerate(file_locs_tobe_uploaded):
            new_loc = file_loc.split('\\')[-1]
            self._signal.emit(f'\t\t{ind}. {new_loc}')
        # 预设一个写文件读取日志的 集合
        files_in_mysql_txt = set()
        for file_loc in file_locs_tobe_uploaded:
            df = pd.DataFrame()
            bad_cols = []
            if task[:5] == 'JP002':
                for st_ind, sheet_name in enumerate([
                        'No Brand', 'No Bullet point', 'No Description', 'Not In Leaf', 'No Title'
                    ]):
                    curr = pd.read_excel(file_loc, sheet_name=sheet_name, skiprows=3)  
                    if st_ind < 4:
                        curr= curr.rename(columns={'Allocation week': "Allocation Week"})
                    ## 检查是否列名不规范 / Check the column names
                    error_msg = f'\nTask {task} -- Cannot find these columns:\n\t '
                    for col in cols:
                        if col not in curr.columns:
                            bad_cols.append(col)
                            error_msg += f"{col}, "
                    if len(bad_cols) > 0:
                        self._bug_signal.emit(error_msg)
                        return False

                    curr = curr.loc[:, cols]
                    df = pd.concat([df, curr])
                    df.reset_index(drop=True, inplace=True)
            else:
                sheet_name = 'History File'
                if task[:5] == 'JP016':
                    skip_row=3
                else:
                    skip_row=1
                try:
                    curr = pd.read_excel(file_loc, sheet_name=sheet_name, skiprows=skip_row)
                except:
                    sheet_name = 'History file'
                    curr = pd.read_excel(file_loc, sheet_name=sheet_name, skiprows=skip_row)
                ## 检查是否列名不规范 / Check the column names
                error_msg = f'\nTask {task} -- Cannot find these columns:\n\t '
                for col in cols:
                    if col not in curr.columns:
                        bad_cols.append(col)
                        error_msg += f"{col}, "
                if len(bad_cols) > 0:
                    self._bug_signal.emit(error_msg)
                    return False

                curr = curr.loc[:, cols]
                df = pd.concat([df, curr])
                df.reset_index(drop=True, inplace=True)
            df['Task'] = task
            df['Year'] = int(file_loc[-9:-5])

            # Filter out un-wanted data in other weeks
            if wanted_all:
                # If the user gave an A to us, we will return data of all weeks
                wks = list(df['Allocation Week'].unique())

                # 因为有些表格第一行数据为示例，可能是“XX"，故需要筛选出数字类型的值
                global all_weeks
                all_weeks = []
                for wk in wks:
                    try:
                        tmp_wk = int(wk)
                    except:
                        print(f"This value for Allocation_week is wrong and will be ignored: {tmp_wk}")
                        continue
                    all_weeks.append(tmp_wk)
                wanted_week = [i for i in all_weeks if i not in in_weeks[file_loc]]      

            df = df.loc[df['Allocation Week'].isin(wanted_week), :]
            if len(df) == 0:
                self._bug_signal.emit(f"{task} -- The data is already in the database!")
            # Update the files_in_mysql if we upload another week of data for this file_loc
            for i in wanted_week:
                in_weeks[file_loc].append(i)
            for k,v in in_weeks.items():
                txt = k + "@"
                for n in v:
                    txt += str(n)
                    txt += " "
                txt += "\n"
                files_in_mysql_txt.add(txt)
            # Remove Nan data
            df = df.dropna()
            # Concatenate dataframes
            all_df = pd.concat([all_df, df])
            all_df.reset_index(drop=True, inplace=True)
        return all_df, files_in_mysql_txt

    def process_read_upload(self, wanted_task, wanted_week, wanted_file, wanted_check):
        ### Read Files
        """
        打开一个 case 的 History file 然后获得所有 CA 关于这个 case 的信息
        Params:
            wanted_task: task name like 'JP010' or 'jp010'
            wanted_week: multiple weeks or one single week
            wanted_file: List[String]
            wanted_check: True/False
        """

        # Input
        task_root = r'\\ant\dept-as\PEK10\DEPT2\RBS\AVS\AVS_Task Related\Proactive Task'
        wanted_columns = ['RBS担当者', 'Allocation Week', 'ASIN', 'RBSステータス']
        task = []
        valid = re.compile("JP\d{3}_")
        for file in os.listdir(task_root):
            if valid.findall(file) and file[:5] == wanted_task:
                # JP*** & 预设任务
                task = file
                break

        f = open('files_in_mysql.txt', 'r')
        global in_files
        in_files = []
        global in_weeks
        in_weeks = {}
        while True:
            fi = f.readline()
            if not fi or len(fi.strip()) == 0:
                break
            if "@" not in fi:
                file_name = fi
                weeks_lst = []
            else:
                try:
                    file_name, weeks = fi.split("@")
                    weeks_lst = [int(i) for i in weeks.strip().split(' ')]
                except:
                    print('Please check the file "files_in_mysql.txt", the data does not fit.')
                    file_name = fi
                    weeks_lst = []
                    continue

            in_files.append(file_name.strip())
            if file_name not in in_weeks:
                in_weeks[file_name] = weeks_lst
            else:
                in_weeks[file_name].extend(weeks_lst)
        f.close()

        ###❤❤❤❤❤ START !!!  为每一个 Task 的各个环节计时
        task_start = time.perf_counter()
        self._signal.emit(f"#💻 Task {task}:")

        #>>> Part 0. Read data from Amazon's Shared Disk
        read_start = time.perf_counter()
        self._signal.emit("\t>>>0. Read data from disk...")
        f_w = open("files_in_mysql.txt", 'w')
        global df
        output = self.read_task_dir(task_root, wanted_columns, task, wanted_week, wanted_file, in_files, in_weeks, f_w)

        # If error happens, return False to stop this process
        if output is False:
            self._bug_signal.emit('\tSomething is wrong, the process is terminated.')
            return False
        
        global files_in_mysql_txt
        df, files_in_mysql_txt = output
        if len(df) == 0:
            self._bug_signal.emit(f"\tNo data to be uploaded")
            return False
        read_end = time.perf_counter()
        self._signal.emit(f"\t\t{round(read_end - read_start, 2)} seconds")

        #>>> Part 1. Identify data and filter out duplicated ones
        self._signal.emit(f"{len(df)} rows of data to be identified")
        if wanted_check:
            filter_start = time.perf_counter()
            self._signal.emit('\t>>>1. Identifying and Filter data(This could take minutes)')
            global record_cnt
            record_cnt = 0
            # We are gonna select all data from Touched
            select_query = """
                SELECT * FROM Touched
            """
            try:
                num = mysql_cursor.execute(select_query)
            except:
                mysql_connect, mysql_cursor = get_mysql_connection()
                num = mysql_cursor.execute(select_query)
            global db_df
            db_df = cursor_into_df(mysql_cursor)
            # 记录已存在的 ASINs,然后对出现过多次的 ASINs,保留 Status 为 Done 的数据，并存入数据库中
            visited_asins = set()
            special_asins = []
            remove_indexes = []
            update_indexes = []
            for ind in tqdm(range(len(df))):
                row = df.iloc[ind]
                login = row['RBS担当者']
                Allocation_week = row['Allocation Week']
                Status = row['RBSステータス']
                Task = row['Task']
                Year = row['Year']
                ASIN = row['ASIN']
                if ASIN in visited_asins:
                    special_asins.append(ASIN)
                else:
                    visited_asins.add(ASIN) 
                # 在提取出的数据库中的数据里，筛选是否有同样的数据
                record = db_df.loc[(db_df.Login == login) & \
                                   (db_df.Allocation_week == Allocation_week) & \
                                   (db_df.Task == Task) & \
                                   (db_df.Allocation_year == Year) & \
                                   (db_df.ASIN == ASIN)]
                if len(record) != 0:
                    # 如果长度=0，说明这个数据可以直接插入数据库，
                    # 如果长度不等于0，意味着该数据已存在，这部分数据应该被更新 -- UPDATE FROM Touched ...
                    update_indexes.append(ind)
                    
                p = 100*(ind+1)/len(df)
                self._p_signal.emit(p)

            # 根据待上传数据中的重复情况，将 Status 不为 Done 的数据移除掉
            for ind, row in df.loc[(df['ASIN'].isin(special_asins)),:].iterrows():
                if row['RBSステータス'] != 'Done':
                    remove_indexes.append(ind)

            global update_df # 用于更新数据库
            update_indexes = [i for i in update_indexes if i not in remove_indexes]
            update_df = df.loc[update_indexes].reset_index(drop=True)
            
            # 剩下的数据是直接上传的
            wanted_index = [i for i in df.index if i not in remove_indexes and i not in update_indexes]
            df = df.iloc[wanted_index,:].drop_duplicates().reset_index(drop=True)

            filter_end = time.perf_counter()
            self._signal.emit(f"\t\t{round(filter_end - filter_start, 2)} seconds")

        else:
            df = df.drop_duplicates().reset_index(drop=True)
            self._signal.emit("\t>>>1. Identify Process is PASSED.")

        #>>> Part 2. Insert new data into Touched if there is any
        self._signal.emit(f"{len(df)} rows of data to be uploaded now.")
        insert_start = time.perf_counter()
        self._signal.emit("\t>>>2. Upload data...")
        time.sleep(1)
        global touched_query        
        touched_query = f"""
            INSERT INTO Touched (Login, Allocation_week, Status, Task, Allocation_year, ASIN)
            VALUES
            """
        if len(df) == 0:
            self._signal.emit(f'All Uploaded before')
        else:
            self._signal.emit("Form the INSERT query for Touched")
            for ind in tqdm(range(len(df)), desc="Form the INSERT query for Touched"):
                row = df.iloc[ind]
                login = row['RBS担当者']
                Allocation_week = row['Allocation Week']
                Status = row['RBSステータス']
                Task = row['Task']
                Year = row['Year']
                ASIN = row['ASIN']
                touched_query += f"""("{login}", {Allocation_week}, "{Status}", "{Task}", {Year}, "{ASIN}"),\n"""
                
                p = 100*(ind+1)/len(df)
                self._p_signal.emit(p)
                
            touched_query = touched_query.strip()[:-1]
            try:
                num = mysql_cursor.execute(touched_query)
                if num > 0:
                    mysql_connect.commit()
                    record_cnt += 1
                elif num != len(df):
                    self._bug_signal.emit("The number of uploaded data is not right.")
                else:
                    print(num, "失败")
                    mysql_connect.rollback()
            except OperationalError:
                # 数据格式有误，比如 Task 的第一行例子
                self._bug_signal.emit(f"{ind}. The data type might not fit the SQL pre-defined one")

            except Exception as e:
                continue_question = input(f"Task {task} has some data stored in this database, do you want to UPDATE instead? \
                            \n(This could cost a bunch of time to check for every row of data)\
                            \nEnter 'Q' to pass the UPDATE process.\n\t")        

                if continue_question.strip().upper() != 'Q':
                    self._signal.emit('You better get prepared, dude!')
                    for ind in tqdm(range(len(df)), desc = f"Updating {task}... "):
                        row = df.iloc[ind]
                        login = row['RBS担当者']
                        Allocation_week = row['Allocation Week']
                        Status = row['RBSステータス']
                        Task = row['Task']
                        Year = row['Year']
                        ASIN = row['ASIN']
                        touched_query = f"""
                        INSERT INTO Touched (Login, Allocation_week, Status, Task, Allocation_year, ASIN)
                        VALUES ("{login}", {Allocation_week}, "{Status}", "{Task}", {Year}, "{ASIN}")
                        """
                        try:
                            num = mysql_cursor.execute(touched_query)
                            if num > 0:
                                mysql_connect.commit()
                                record_cnt += 1
                            else:
                                mysql_connect.rollback()

                        except IntegrityError:
                            # Status 更新
                            update_sql = f"""
                                UPDATE Touched SET Status ="{Status}" 
                                WHERE Login = "{login}" AND Allocation_week = {Allocation_week}
                                AND Task = "{Task}" AND Allocation_year = {Year} AND ASIN = "{ASIN}"
                            """
                            try:
                                mysql_cursor.execute(update_sql)
                                mysql_connect.commit()
                            except Exception as e:
                                bug_msg = str(e) + 'Failed to append data'
                                self._bug_signal.emit(bug_msg)  
                        except OperationalError:
                            # 有的 Task 有一行例子
                            continue       
            insert_end = time.perf_counter()
            self._signal.emit(f"\t{round(insert_end - insert_start, 2)} seconds")

        if wanted_check:
            # 跨表更新 Touched 效率高一些
            # 建立 Temp_touched 表，用于上传更新后便删除，常态下，该表长度为0
            #>>> Part 3. Feed uploaded data into a temporary database
            self._signal.emit('\t>>3. Feed duplicated data into Temp_touched...')
            temp_start = time.perf_counter()
            time.sleep(1)
            global temp_query        
            temp_query = f"""
                INSERT INTO Temp_touched (Login, Allocation_week, Status, Task, Allocation_year, ASIN)
                VALUES\n"""
            update_df = update_df.drop_duplicates().reset_index(drop=True)
            for ind in tqdm(range(len(update_df)), desc=f"Form the query for Temp_touched"):
                row = update_df.iloc[ind]
                login = row['RBS担当者']
                Allocation_week = row['Allocation Week']
                Status = row['RBSステータス']
                Task = row['Task']
                Year = row['Year']
                ASIN = row['ASIN']
                temp_query += f"""("{login}", {Allocation_week}, "{Status}", "{Task}", {Year}, "{ASIN}"),\n"""
                
                p = 100*(ind+1)/len(update_df)
                self._p_signal.emit(p)
                
            temp_query = temp_query.strip()[:-1]   
            try:
                mysql_cursor.execute(temp_query)
            except Exception as e:
                self._bug_signal.emit(str(e))
                self._signal.emit('\t( :C Probably, the Temp_touched was not cleared last time)')
                # 清空 临时库数据
                rm_query = """DELETE FROM Temp_touched"""
                mysql_cursor.execute(rm_query)
                mysql_connect.commit()
                time.sleep(1)
                mysql_cursor.execute(temp_query)
            mysql_connect.commit()
            temp_end = time.perf_counter()
            print(f"\t{round(temp_end - temp_start, 2)} seconds")

            #>>> Part 4. Update the Touched
            update_start = time.perf_counter()
            self._signal.emit('\t>>4. Begin to update database...')
            global update_query
            update_query = f"""
                UPDATE Touched 
                INNER JOIN Temp_touched t
                ON Touched.Login = t.Login 
                    AND Touched.Allocation_week = t.Allocation_week
                    AND Touched.Task = t.Task
                    AND Touched.Allocation_year = t.Allocation_year 
                    AND Touched.ASIN = t.ASIN
                SET Touched.Status = t.Status
            """
            mysql_cursor.execute(update_query)
            mysql_connect.commit()
            update_end = time.perf_counter()
            self._signal.emit(f"\t{round(update_end - update_start, 2)} seconds")
        else:
            self._signal.emit("\t>>>4. Update Process is PASSED.")

        #>>> Part 5. 整理读写日志 以及 清空临时表
        empty_start = time.perf_counter()
        self._signal.emit("\t>>5. Compile files_in_mysql.txt and Empty the temporary table...")

        for txt in files_in_mysql_txt:
            f_w.write(txt)
        f_w.close()

        rm_query = """DELETE FROM Temp_touched"""
        empty_start = time.perf_counter()
        mysql_cursor.execute(rm_query)
        mysql_connect.commit() 
        empty_end = time.perf_counter()
        self._signal.emit(f'\t{round(empty_end-empty_start, 2)} seconds')


        ###❤❤❤❤❤ END !!!  
        task_end = time.perf_counter()
        self._signal.emit(f'\n\t♥ Successful! It took us {round(task_end - task_start, 2)} seconds for this task.')
    

class Thread_Upload_CM(QThread):
    _signal = pyqtSignal(str)
    _bug_signal = pyqtSignal(str) #self._bug_signal.emit()
    _p_signal = pyqtSignal(int) # 进度信号
    def __init__(self):
        super().__init__()
        
    def run(self):
        
        curr = datetime.datetime.now().strftime("%Y/%m/%d")
        curr_week = int(time.strftime("%W").strip()) # 今年没有第一周
        curr_year = int(time.strftime("%Y").strip())

        # 根文件夹位置
        root = r"\\ant\dept-as\PEK10\DEPT2\RBS\AVS\AVS_Task Related\Proactive Task\[Updating] Proactive work allocation\Work Allocation File"
        cm_data = root + "\\CM Data"
        
        # 根据文件夹内的 xlsx 文件 获取当前 CM 名单
        try:
            Relation_tb = pd.DataFrame(columns=["CM","CA"])
            Allocation_tb = pd.DataFrame(columns=["CA","Task","Base","Range","Week_from", "Week_to"])
        except:
            self._bug_signal.emit(f"Please check the column names of {cm_data} before your next try.")
            return
        
        CMs = []
        for file in os.listdir(cm_data):
            if "~$" in file:
                continue
            name = file.split(' team')[0]
            CMs.append(name)
            week_from, week_to = file[:-5].split('_wk')[1].split('-')

            if int(week_from) > curr_week:
                file_year = curr_year - 1
            else:
                file_year = curr_year

            file_path = cm_data + "\\" + file

            rela, allo = self.cm_file(file_path, name)

            allo['Week_from'] = week_from
            allo['Week_to'] = week_to
            allo['Year'] = file_year

            Relation_tb = pd.concat([Relation_tb, rela], axis=0)
            Allocation_tb = pd.concat([Allocation_tb, allo], axis=0)

            Relation_tb.reset_index(drop=True,inplace=True)
            Allocation_tb.reset_index(drop=True,inplace=True)
        
        # 上传数据至 Management
        Relation_tb = Relation_tb.drop_duplicates().reset_index(drop=True)
        self.UpdateMgmt(Relation_tb)    
        
        # 上传数据至 Allocation
        Allocation_tb.Range = Allocation_tb.Range.astype('float')
        Allocation_tb.Week_from = Allocation_tb.Week_from.astype('int')
        Allocation_tb.Week_to = Allocation_tb.Week_to.astype('int')
        Allocation_tb.Year = Allocation_tb.Year.astype('int')
        self.UpdateAllo(Allocation_tb)
        
        
        
    def UpdateMgmt(self, Relation_tb):
        mysql_connect, mysql_cursor = get_mysql_connection()
        # 将 CA-CM 写入数据库 Management
        self._signal.emit("[2/3] Upload data into Management table")

        # 1. Check if duplicated or needs update
        update_index = []
        insert_index = []
        self._signal.emit(f"{len(Relation_tb)} rows of data needs identification.")
        for ind in tqdm(range(len(Relation_tb))):
            row = Relation_tb.iloc[ind]
            CM = row['CM']
            CA = row['CA']
            sql = f"""SELECT * FROM Management WHERE CA = "{CA}" """
            try:
                mysql_cursor.execute(sql)
            except:
                mysql_connect, mysql_cursor = get_mysql_connection()
                mysql_cursor.execute(sql)
            df = cursor_into_df(mysql_cursor)
            if len(df) == 1 and df.loc[0, 'CM'] != CM:
                update_index.append(ind)
            elif len(df) == 0:
                insert_index.append(ind)
                
            p = 100*(ind+1)/len(Relation_tb)
            self._p_signal.emit(p)


        # 1.5 将需要 Update 的数据抽出来并更新
        if len(update_index) > 0:
            update_df = Relation_tb.loc[update_index].reset_index(drop=True)
            self._signal.emit(f"{len(update_df)} rows of data will be updated within Management")
            for ind in tqdm(range(len(update_df))):
                row = update_df.loc[ind]
                CM = row['CM']
                CA = row['CA']
                update_sql = f"""
                    UPDATE Management SET CM ="{CM}" WHERE CA = "{CA}"
                """
                try:
                    mysql_cursor.execute(update_sql)
                    mysql_connect.commit()
                except Exception as e:
                    print(e,'数据库修改失败')  
                    
                p = 100*(ind+1)/len(update_df)
                self._p_signal.emit(p)
        else:
            self._signal.emit("Management表 -- 无需更新")

        # 2.0 直接 Insert
        insert_df = Relation_tb.loc[insert_index].reset_index(drop=True)
        sql = f"""INSERT INTO Management (CM, CA) VALUES """

        if len(insert_df) != 0:
            self._signal.emit(f"{len(insert_df)} rows of data will be inserted into Management")
            for ind in tqdm(range(len(insert_df))):
                row = insert_df.loc[ind]
                CM = row['CM']
                CA = row['CA']
                sql += f"""("{CM}", "{CA}"),\n"""
            sql = sql.strip()[:-1]  

            try:
                num = mysql_cursor.execute(sql)
                if num > 0:
                    mysql_connect.commit()
                else:
                    mysql_connect.rollback()
            except Exception as e:
                self._bug_signal.emit(str(e))
        else:
            self._signal.emit("Management表 -- 无新数据")

        self._signal.emit('Management表 -- 数据更新完毕')
        
    def UpdateAllo(self, Allocation_tb):
        mysql_connect, mysql_cursor = get_mysql_connection()
        # 写入 Allocation
        self._signal.emit("[3/3] Upload data into Allocation table")
        Allocation_tb = Allocation_tb.dropna().reset_index(drop=True)

        # 1. Check if duplicated or needs update
        update_index = []
        insert_index = []
        self._signal.emit(f"{len(Allocation_tb)} rows of data needs identification.")
        for ind in tqdm(range(len(Allocation_tb))):
            row = Allocation_tb.loc[ind]
            sql = f"""
                SELECT * FROM Allocation
                WHERE CA = "{row.CA}" AND Task = "{row.Task}" 
                AND Week_from = {row.Week_from} AND Week_to = {row.Week_to}
                AND Year = {row.Year} 
            """
            mysql_cursor.execute(sql)
            df = cursor_into_df(mysql_cursor)
            if len(df) == 1 and (df.loc[0, 'Basic_val'] != row.Base or df.loc[0, 'Range_val'] != row.Range):
                update_index.append(ind)
            elif len(df) == 0:
                insert_index.append(ind)
                
            p = 100*(ind+1)/len(Allocation_tb)
            self._p_signal.emit(p)
                
        # 1.5 将需要 Update 的数据抽出来并更新
        if len(update_index) > 0:
            update_df = Allocation_tb.loc[update_index].reset_index(drop=True)
            self._signal.emit(f"{len(update_df)} rows of data will be updated within Allocation")
            for ind in tqdm(range(len(update_df))):
                row = update_df.loc[ind]
                update_sql = f"""
                    UPDATE Allocation SET Basic_val = {row.Base}, Range_val = {row.Range} 
                    WHERE CA = "{row.CA}" AND Task = "{row.Task}" 
                    AND Week_from = {row.Week_from} AND Week_to = {row.Week_to}
                    AND Year = {row.Year} 
                """
                try:
                    mysql_cursor.execute(update_sql)
                    mysql_connect.commit()
                except Exception as e:
                    self._bug_signal.emit(e,'数据库修改失败')    
                    
                p = 100*(ind+1)/len(update_df)
                self._p_signal.emit(p)
            
        else:
            self._signal.emit("Allocation表 -- 无需更新")
            
        # 2.0 直接 Insert
        global insert_df, insert_allo_sql
        if len(insert_index) > 0:
            insert_df = Allocation_tb.loc[insert_index].reset_index(drop=True)
            insert_allo_sql = f"""INSERT INTO Allocation (CA, Task, Basic_val, Range_val, Week_from, Week_to, Year) VALUES \n"""
            self._signal.emit(f"{len(insert_df)} rows of data will be inserted into Allocation")
            for ind in tqdm(range(len(insert_df))):
                row = insert_df.loc[ind]
                insert_allo_sql += f"""("{row.CA}", "{row.Task}", {row.Base}, {row.Range}, {row.Week_from}, {row.Week_to}, {row.Year}), \n"""

            insert_allo_sql = insert_allo_sql.strip()[:-1]  

            try:
                num = mysql_cursor.execute(insert_allo_sql)
                if num > 0:
                    mysql_connect.commit()
                else:
                    mysql_connect.rollback()
            except Exception as e:
                self._bug_signal.emit(str(e))
        else:
            self._signal.emit("Allocation表 -- 无新数据")
            
        self._signal.emit('Allocation表 -- 数据更新完毕')
        
    def cm_file(self, file_path, CM):
        """
        params:
            file_path: CM 放置在公盘里的excel的路径
            CM: Catalog Manager 的 名字，方便存档
        returns:
            Allocations: 储存 CA - Task - Base - Range
            Relations : 储存 CM - CA
        """

        base = pd.read_excel(file_path, sheet_name="Base")
        rage = pd.read_excel(file_path, sheet_name="Range")

        Relations = pd.DataFrame(columns=["CM","CA"])   
        Allocations = pd.DataFrame(columns=["CA","Task","Base","Range"]) 

        for ind in range(2, len(base.columns)):
            CA = base.columns[ind]
            relation = {"CM":CM, "CA":CA}
            Relations = Relations.append(relation, ignore_index=True)
            for row in range(len(base)-1):
                task = base.iloc[row,0]
                b = base.iloc[row,ind]
                r = rage.iloc[row,ind]
                kid = {'CA': CA, 'Task':task, 'Base':b, "Range":r}
                Allocations = Allocations.append(kid, ignore_index=True)

        rela_col = ["CM","CA"]
        rela_col_dict = {ind: rela_col[ind] for ind in range(len(rela_col))}
        Relations.rename(rela_col_dict, inplace=True, axis=1)

        allo_col = ["CA","Task","Base","Range"]
        allo_col_dict = {ind: allo_col[ind] for ind in range(len(allo_col))}
        Allocations.rename(allo_col_dict, inplace=True, axis=1)

        Relations.reset_index(drop=True,inplace=True)
        Allocations.reset_index(drop=True,inplace=True)

        return Relations, Allocations

    
class Thread_get_data(QThread):
    _signal =pyqtSignal(int)
    html_signal = pyqtSignal(str)
    text_signal = pyqtSignal(str)
    single_input_signal = pyqtSignal(list)
    multiple_input_signal = pyqtSignal(list)
    def __init__(self, datas):
        super().__init__()
        self.datas = datas
        self.return_datas = None
        
    def run(self):
        qmut_init.lock()
        qmut_single.lock()
        qmut_multiple.lock()
        self._signal.emit(0)
        q_wk_from, q_wk_to, q_year, weeks, years, the_CM, CM_login, the_CMs, CMs_login, \
                       task_list_dir, not_wanted_status = self.datas
        
        # 连接数据库
        connected_msg_html = """<font color="#DD7907">Begin to connect to the database</font> """
        print("Begin to connect to the database")
        self.html_signal.emit(connected_msg_html)
        # 1.2 Connect to DB
        # 1.2.1 连接 MySQL 数据库并写入数据                
        avs_pro_connect = pymysql.connect(host='dev-dsk-wangting-2a-25bc2431.us-west-2.amazon.com', 
                                  user='avs_user', 
                                  password='avs_pro', 
                                  database='AVS_proactive',
                                  charset='utf8') #服务器名,账户,密码,数据库名
        avs_pro_cursor = avs_pro_connect.cursor()
        self._signal.emit(30)
        # 1.2.2 连接 MySQL 数据库，获取 TWT 数据
        twt_connect = pymysql.connect(host='dev-dsk-wangting-2a-25bc2431.us-west-2.amazon.com', 
                                  user='TWT_USER', 
                                  password='TWT_USER', 
                                  database='TWT',
                                  charset='utf8') #服务器名,账户,密码,数据库名
        twt_cursor = twt_connect.cursor()
        self._signal.emit(40)
        # 1.2.3 连接 RedShift 数据库
        rs_connect = psycopg2.connect(database="rsbidw",
                                user="avs_user",
                                password="A21sP22oR$",
                                host="rsbi-analytics.clszsz7jap6y.us-east-1.redshift.amazonaws.com",
                                port="8192")
        rs_cursor = rs_connect.cursor()
        self._signal.emit(50)
        connects = [rs_connect, avs_pro_cursor, twt_cursor]
        connected_msg_html = """<font color="#FF0000">Connected!</font> """
        self.html_signal.emit(connected_msg_html)
        print('Connected')
        
        ## 创建 HTML 和 Data 的存放文件夹
        folder_created_html = """<font color="#DD7907">Check and create HTML & DataFrame folders</font> """
        self.html_signal.emit(folder_created_html)
        print("Check and create HTML & DataFrame folders")
        b = os.getcwd()
        html_dir = b + "\\HTMLs"
        data_dir = b + "\\DataFrames"
        if not os.path.exists(html_dir):
            os.mkdir(html_dir)
        if not os.path.exists(data_dir):
            os.mkdir(data_dir)
            
        self.data_dir = data_dir
        self.data_report_dir = data_dir
        self.html_dir = html_dir
        folder_created_html = """<font color="#FF0000">Created!</font> """
        self.html_signal.emit(folder_created_html)
        print("Created!")
        
        # GET DATA
        get_data_msg_html = """<font color="#DD7907">Procedure is on to analyze data</font> """
        self.html_signal.emit(get_data_msg_html)
        print("Procedure is on to analyze data")
        single_week_data = get_data(avs_pro_cursor, twt_cursor, rs_cursor, 
                             q_wk_from, q_wk_to, q_year,
                             the_CM, CM_login, task_list_dir, not_wanted_status)
        self._signal.emit(70)
        multiple_weeks_data = get_by_task_data(avs_pro_cursor, twt_cursor, rs_cursor,
                                weeks, years, 
                                the_CMs, CMs_login, task_list_dir, not_wanted_status)
        get_data_msg_html = """<font color="#FF0000">Ready to create Reports!</font> """
        self.html_signal.emit(get_data_msg_html)
        print("Ready to create Reports!")
        self._signal.emit(100)
        
        self.return_datas = [html_dir, data_dir, connects, single_week_data, multiple_weeks_data]
        
        qmut_init.unlock()
        qmut_single.unlock()
        qmut_multiple.unlock()
        
    
class Thread_single_week(QThread):
    _signal = pyqtSignal(int)
    _name_signal = pyqtSignal(str) # 用于发送 html 的文件位置
    def __init__(self, datas):
        super().__init__()
        self.datas = datas
        self.return_datas = None
        
    def run(self):
        qmut_single.lock()
        self._signal.emit(0)
        time.sleep(0.2)
        q_wk_from, q_wk_to, q_year, \
            weeks, years, \
            the_CM, CM_login, \
            the_CMs, CMs_login, \
            task_list_dir, \
            data_dir, html_dir, \
            single_week_data, viz_params_time_compare, viz_params_by_person, \
            tb_names, tb_subtitles, pb_names, pb_subtitles, intro_3, intro_10 = self.datas
        touched_df, allocation_df, allo_touched, atwt, atwt_wp, case_cnt = single_week_data 
        # INIT
        ignore_params, default_selected, Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, \
            row_sheet_counts, total_metrics, title_padding = viz_params_time_compare
        total = total_metrics - ignore_params
        ## 其余参数通过计算得到
        col_sheet_counts, top_loc, bottom_loc, left_loc, right_loc, Grid_height = \
            get_params(Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, total, ignore_params)
        hundred_percent_line = """opts.MarkLineItem(name="100%%", y=1, symbol='diamond', symbol_size=[10,10])"""
        tb_mark_100_line = [
            "", "", "", "","",
        ]
        ##  HTML 内容 
        locs = {}
        glb = {"Bar": Bar, 'opts':opts}
        c1 = f"""c = (Bar(
                    init_opts=opts.InitOpts(
                        bg_color = "#FFFFFF"
                    )
                )
            .add_xaxis(CA_names)\n"""
        c2 = f"""
            .set_series_opts(
                label_opts=opts.LabelOpts(is_show=False),
                markpoint_opts=opts.MarkPointOpts(
                    data=[
                        opts.MarkPointItem(type_="max", name="最大值"),
                    ],
                    symbol="circle",
                    symbol_size=[1,1],
                    label_opts=opts.LabelOpts(
                        position="top", color="#0A0A0D",
                        font_size=9, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

                markline_opts=opts.MarkLineOpts(
                    data=[
                        opts.MarkLineItem(type_="average", name="平均值"),
                        %s
                    ],
                    linestyle_opts = opts.LineStyleOpts(
                        opacity=0.5,
                        width=0.6,
                        type_="dotted",
                    ),
                    symbol_size=[0,1],
                    label_opts=opts.LabelOpts(
                        position="right", 
                        font_size=10, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

            )
            .set_global_opts(
                xaxis_opts=opts.AxisOpts(
                    axislabel_opts=opts.LabelOpts(rotate=0),
                    name='CA'
                ),
                yaxis_opts=opts.AxisOpts(
                    name=''
                ),
                title_opts=opts.TitleOpts(title="%s", subtitle="%s", 
                    item_gap=5,
                    pos_top=f"%d%%",
                    pos_bottom=f"%d%%",
                    pos_left=f"%d%%",
                    pos_right=f"%d%%",  
                    title_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    subtitle_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    %s
                ),
                brush_opts=opts.BrushOpts(),
                datazoom_opts=opts.DataZoomOpts(
                    type_="inside",
                    pos_bottom = "10%%",
                    filter_mode="empty"
                ),
                toolbox_opts=opts.ToolboxOpts(
                        pos_left = "40%%",
                        pos_top = "0%%",
                    ),
                legend_opts=opts.LegendOpts(
                        type_ = 'scroll',
                        selected_mode = 'multiple',
                        orient = 'vertical',
                        pos_right = '1%%',
                        pos_left = '{103-right_loc[ignore_params+col_sheet_counts-1]}%%',
                        pos_top = '10%%',
                        align = 'right',
                        item_gap = 20,
                        padding = 5,
                        textstyle_opts = opts.TextStyleOpts(font_size = %d)
                    ),
                tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),

            ))"""
        cs = [None for i in range(total)] ## 存放所有打算合并的 HTML
        
        ##########     1
        global CA_names
        tb_1_df = atwt_wp \
                .groupby(['CA', 'Task']).sum() \
                .asins_updated.reset_index()

        tb_1_df_new = enrich_df_tasks(tb_1_df)
        tb_1_df_new = tb_1_df_new.groupby(['CA', 'Task']).sum().reset_index().sort_values(by=['CA', 'Task'])
        tb_tasks =list(tb_1_df_new.Task.unique())
        CA_names = list(tb_1_df.CA.unique())
        glb['CA_names'] = CA_names
        tb_ind = 0
        curr_c1 = "" + c1
        completed_asins_lists = [None for i in tb_tasks]
        # Generate Viz
        for ind, task in enumerate(tb_tasks):
            completed_asins_lists[ind] = list(tb_1_df_new.loc[tb_1_df_new.Task == task,:].asins_updated.values)
            completed_asins_lists[ind] = [int(i) for i in completed_asins_lists[ind]]
            curr_c1 += f""".add_yaxis("{task[:5]}", {completed_asins_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                      3, 95, 10, 90, 
                      20, 15, 
                      "",10)
        glb['completed_asins_lists'] = completed_asins_lists
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\Time_Compare__{tb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{tb_names[tb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if tb_ind >= ignore_params:
            hyperlink = fr"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                          top_loc[tb_ind]-(title_padding*grid_vert), bottom_loc[tb_ind], 
                          left_loc[tb_ind], right_loc[tb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[tb_ind-ignore_params]=c

        ## 数据表格
        completed_asins_df = pd.DataFrame(columns = CA_names, index=tb_tasks)
        iii=0
        for ind, row in completed_asins_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                completed_asins_df.loc[ind, ca] = completed_asins_lists[iii][ca_ind]
            iii += 1
            
        self._signal.emit(8)
        time.sleep(0.2)
        ##########     2
        tb_2_df = atwt_wp.loc[atwt_wp.CM == the_CM, :] \
                .groupby(['CA', 'Task']).sum() \
                .Hour.reset_index()
        tb_2_df_new = enrich_df_tasks(tb_2_df, tb_tasks)
        tb_2_df_new = tb_2_df_new.groupby(['CA', 'Task']).sum().reset_index().sort_values(by=['CA', 'Task'])

        tb_ind = 1
        curr_c1 = "" + c1
        time_lists = [None for i in tb_tasks]
       
        # Generate Viz    
        for ind, task in enumerate(tb_tasks):
            time_lists[ind] = list(tb_2_df_new.loc[tb_2_df_new.Task == task,:].Hour.values)
            time_lists[ind] = [round(i,2) for i in time_lists[ind]]
            curr_c1 += f""".add_yaxis("{task[:5]}", {time_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""
        glb['time_lists'] = time_lists
        curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                      3, 95, 10, 90,
                      20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\Time_Compare__{tb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{tb_names[tb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if tb_ind >= ignore_params:
            hyperlink = fr"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                          top_loc[tb_ind]-(title_padding*grid_vert), bottom_loc[tb_ind], 
                          left_loc[tb_ind], right_loc[tb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[tb_ind-ignore_params]=c

        ## 数据表格
        time_df = pd.DataFrame(columns = CA_names, index=tb_tasks)
        iii=0
        for ind, row in time_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                time_df.loc[ind, ca] = time_lists[iii][ca_ind]
            iii += 1
        self._signal.emit(13)
        time.sleep(0.2)
        ##########     3
        tb_3_df = enrich_df_tasks(case_cnt, tb_tasks)
        tb_3_df = tb_3_df.sort_values(by=['CA', 'Task']).fillna(value=0)
        tb_ind = 2
        curr_c1 = "" + c1
        cases_cnt_lists = [None for i in tb_tasks]
        asin_per_case_lists = [None for i in tb_tasks]
        
        # Generate Viz
        for ind, task in enumerate(tb_tasks):
            tmp = list(tb_3_df.loc[tb_3_df.Task == task,:].Cases_count.values)
            cases_cnt_lists[ind] = tmp
            asin_per_case_lists[ind] = [round(completed_asins_lists[ind][i]/tmp[i], 2) if tmp[i] !=0 else 0 for i in range(len(tmp))]
            curr_c1 += f""".add_yaxis("{task[:5]}", {asin_per_case_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""
        glb['asin_per_case_lists'] = asin_per_case_lists
        glb['cases_cnt_lists'] = cases_cnt_lists
        curr_c2 = c2%(tb_mark_100_line[tb_ind], 
                      tb_names[tb_ind], tb_subtitles[tb_ind], 
                      3, 95, 10, 90, 
                      20, 15,
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\Time_Compare__{tb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{tb_names[tb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if tb_ind >= ignore_params:
            hyperlink = fr"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                          top_loc[tb_ind]-(title_padding*grid_vert), bottom_loc[tb_ind], 
                          left_loc[tb_ind], right_loc[tb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[tb_ind-ignore_params]=c

        ## 数据表格
        asin_per_case_df = pd.DataFrame(columns = CA_names, index=tb_tasks)
        iii=0
        for ind, row in asin_per_case_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                asin_per_case_df.loc[ind ,ca] = asin_per_case_lists[iii][ca_ind]
            iii += 1
            
        cases_df = pd.DataFrame(columns = CA_names, index=tb_tasks)
        iii = 0
        for ind, row in cases_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                cases_df.loc[ind, ca] = cases_cnt_lists[iii][ca_ind]
            iii+=1
        self._signal.emit(21)
        time.sleep(0.2)
        ##########     4
        tb_ind = 3
        cases_per_hour_lists = [None for i in tb_tasks]
        # Generate Viz   
        curr_c1 = "" + c1
        for ind, task in enumerate(tb_tasks):
            tmp = []
            for i in range(len(cases_cnt_lists[ind])):
                if time_lists[ind][i]!=0:
                    val = round(cases_cnt_lists[ind][i]/time_lists[ind][i], 2)
                else:
                    val = 0
                tmp.append(val)
            cases_per_hour_lists[ind] = tmp
            curr_c1 += f""".add_yaxis("{task[:5]}", {cases_per_hour_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""
        glb['cases_per_hour_lists'] = cases_per_hour_lists
        curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                      3, 95, 10, 90,
                      20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\Time_Compare__{tb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{tb_names[tb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if tb_ind >= ignore_params:
            hyperlink = fr"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                          top_loc[tb_ind]-(title_padding*grid_vert), bottom_loc[tb_ind], 
                          left_loc[tb_ind], right_loc[tb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[tb_ind-ignore_params]=c

        ## 数据表格
        cases_per_hour_df = pd.DataFrame(columns = CA_names, index=tb_tasks)
        iii=0
        for ind, row in cases_per_hour_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                cases_per_hour_df.loc[ind, ca] = cases_per_hour_lists[iii][ca_ind]
            iii += 1
        self._signal.emit(27)
        time.sleep(0.2)
        
        ##########     5
        tb_ind = 4
        asin_per_hour_lists = [None for i in tb_tasks]
        # Generate Viz
        curr_c1 = "" + c1
        for ind, task in enumerate(tb_tasks):
            l, r = completed_asins_lists[ind], time_lists[ind]
            asin_per_hour_lists[ind] = [round(l[i]/r[i], 2) if r[i] !=0 else 0 for i in range(len(CA_names))]
            curr_c1 += f""".add_yaxis("{task[:5]}", {asin_per_hour_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        glb['asin_per_hour_lists'] = asin_per_hour_lists
        c2_patch = "" + c2
        curr_c2 = c2_patch % (tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                        3, 95, 10, 90, 
                        20, 12, 
                        "", 10)
        exec(curr_c1 + curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\Time_Compare__{tb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{tb_names[tb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)
        if tb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2_patch%(tb_mark_100_line[tb_ind], tb_names[tb_ind], tb_subtitles[tb_ind], 
                          top_loc[tb_ind]-(title_padding*grid_vert), bottom_loc[tb_ind], 
                          left_loc[tb_ind], right_loc[tb_ind], 
                          15, 12,
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[tb_ind-ignore_params]=c

        ## 数据表格
        asin_per_hour_df = pd.DataFrame(columns=CA_names, index=tb_tasks)
        iii = 0
        for ind, row in asin_per_hour_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                asin_per_hour_df.loc[ind, ca] = asin_per_hour_lists[iii][ca_ind]
            iii += 1
        self._signal.emit(34)
        ########### Grid All Sheets
        time_compare_grid = Grid(
            init_opts=opts.InitOpts(width=f"{Grid_width}px",height=f"{Grid_height}px", bg_color = "#FFFFFF"))
        for ind in range(total):
            c = cs[ind]
            time_compare_grid \
                .add(c,
                     grid_opts=opts.GridOpts(
                         pos_top=f"{top_loc[ind+ignore_params]}%",
                         pos_bottom=f"{bottom_loc[ind+ignore_params]}%",
                         pos_left=f"{left_loc[ind+ignore_params]}%",
                         pos_right=f"{right_loc[ind+ignore_params]}%",
                     ),
                is_control_axis_index=False
                )
        html_name = f"Time_Compare_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}.html"
        time_compare_grid.render(html_name)
        self.time_compare_grid = time_compare_grid
        self._name_signal.emit(html_name)
        
        ## Save Data Frame
        # 多加一行 raw data,所以重新定义 tb_names
        tb_names = [
            '# Asins',
            '# Hours',
            '# Cases', 
            'Asins per Case',
            'Cases per Hour',
            'Asins per Hour'
        ]
        time_compare_dfs = [
            completed_asins_df, time_df, cases_df,
            asin_per_case_df, cases_per_hour_df, asin_per_hour_df
        ]
        all_sheets = []
        excel_file_loc = fr"{data_dir}\\Time Compare REPORT_{q_wk_from}_{q_wk_to}_{q_year}_{the_CM}.xlsx"
        with pd.ExcelWriter(excel_file_loc) as xlsx:
            for task in tb_tasks:
                TimeCompare_df = pd.DataFrame()
                for df_ind, df in enumerate(time_compare_dfs):
                    df['Metric'] = tb_names[df_ind]
                    df['indexes'] = df_ind
                    curr = pd.DataFrame(df.loc[task,:]).T
                    TimeCompare_df = pd.concat([TimeCompare_df, curr])
                output = TimeCompare_df.groupby(['Metric']).sum().sort_values('indexes').iloc[:, :-1]
                output.to_excel(xlsx, sheet_name=task[:5], index=True)
            

        revise_workbook(excel_file_loc, CA_names, tb_names, all_sheets)
        # add Explanation
        wb = load_workbook(excel_file_loc)
        intro_sheet_name = 'Intro'
        if intro_sheet_name not in wb.sheetnames:
            # 创建 Sheet 并塞入内容
            ws = wb.create_sheet(intro_sheet_name, 0)
            ws['A1'] = 'Metrics'
            ws['B1'] = 'Intro'
            ws['A2'] = '# Asins'
            ws['A3'] = '# Hours'
            ws['A4'] = '# Cases'
            for i in range(len(intro_3)):
                ws[f'B{i+2}'] = intro_3[i]

            # 设置行高&列宽
            #第一行20,后面3行18;第一列10,第二列30
            ws.row_dimensions[1].height=20
            for row in range(2,5):
                ws.row_dimensions[row].height = 18
            ws.column_dimensions['A'].width=10
            ws.column_dimensions['B'].width=30

            # 字体格式
            #边框
            thin = Side(border_style="thin", color="000000")#边框样式，颜色
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
            #字体
            init_font = Font(size=10, bold=True, name='Microsoft YaHei',  color="101013")
            cell_font = Font(size=10, bold=False, name='Microsoft YaHei',  color="101013")
            #居中
            cell_align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            for row in ws['A1:B4']:
                for cell in row:
                    if cell.column_letter == 'A' or cell.row == 1:
                        cell.font = init_font
                    else:
                        cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = cell_border
            # 底色
            for letter in ['A', 'B']:
                ws[f'{letter}1'].fill = fill_1
            wb.save(excel_file_loc)
        print("Time Compare Report is created!")
        self._signal.emit(47)
        time.sleep(0.5)
        
        
        
        ########################### By Person
        # INIT
        ignore_params, default_selected, Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, \
            row_sheet_counts, total_metrics, title_padding = viz_params_by_person
        total = total_metrics - ignore_params
        ## 其余参数通过计算得到
        col_sheet_counts, top_loc, bottom_loc, left_loc, right_loc, Grid_height = \
            get_params(Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, total, ignore_params)

        hundred_percent_line = """opts.MarkLineItem(name="100%%", y=1, symbol='diamond', symbol_size=[10,10])"""
        bp_mark_100_line = [
            "", "", "", "",
            hundred_percent_line,
            hundred_percent_line,
            "", "", "", 
            hundred_percent_line
        ]

        c1 = f"""c = (Bar(
                    init_opts=opts.InitOpts(
                        bg_color = "#FFFFFF"
                    )
                )
            .add_xaxis(CA_names)\n"""
        c2 = f"""
            .set_series_opts(
                label_opts=opts.LabelOpts(is_show=False),
                markpoint_opts=opts.MarkPointOpts(
                    data=[
                        opts.MarkPointItem(type_="max", name="最大值"),
                    ],
                    symbol="circle",
                    symbol_size=[1,1],
                    label_opts=opts.LabelOpts(
                        position="top", color="#0A0A0D",
                        font_size=9, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

                markline_opts=opts.MarkLineOpts(
                    data=[
                        opts.MarkLineItem(type_="average", name="平均值"),
                        %s
                    ],
                    linestyle_opts = opts.LineStyleOpts(
                        opacity=0.5,
                        width=0.6,
                        type_="dotted",
                    ),
                    symbol_size=[0,1],
                    label_opts=opts.LabelOpts(
                        position="right", 
                        font_size=10, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

            )
            .set_global_opts(
                xaxis_opts=opts.AxisOpts(
                    axislabel_opts=opts.LabelOpts(rotate=0),
                    name='CA'
                ),
                yaxis_opts=opts.AxisOpts(
                    name=''
                ),
                title_opts=opts.TitleOpts(title="%s", subtitle="%s", 
                    item_gap=5,
                    pos_top=f"%d%%",
                    pos_bottom=f"%d%%",
                    pos_left=f"%d%%",
                    pos_right=f"%d%%",  
                    title_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    subtitle_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    %s
                ),
                brush_opts=opts.BrushOpts(),
                datazoom_opts=opts.DataZoomOpts(
                    type_="inside",
                    pos_bottom = "10%%",
                    filter_mode="empty"
                ),
                toolbox_opts=opts.ToolboxOpts(
                        pos_left = "40%%",
                        pos_top = "0%%",
                    ),
                legend_opts=opts.LegendOpts(
                        type_ = 'scroll',
                        selected_mode = 'multiple',
                        orient = 'vertical',
                        pos_right = '1%%',
                        pos_left = '{103-right_loc[ignore_params+col_sheet_counts-1]}%%',
                        pos_top = '10%%',
                        align = 'right',
                        item_gap = 20,
                        padding = 5,
                        textstyle_opts = opts.TextStyleOpts(font_size = %d)
                    ),
                tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),

            ))"""
        cs = [None for i in range(total)]
        
        ##########     1
        # Generate Viz  
        global pb_tasks
        pb_ind = 0
        pb_bg = enrich_df_tasks(allocation_df)
        pb_tasks = list(pb_bg.Task.unique())
        curr_c1 = "" + c1
        basic_goal_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            basic_goal_lists[ind] = list(pb_bg.loc[pb_bg.Task == task,:].Basic_val.values)
            basic_goal_lists[ind] = [round(i,0) for i in basic_goal_lists[ind]]
            curr_c1 += f""".add_yaxis("{task[:5]}", {basic_goal_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 
                      20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],
                          pb_names[pb_ind], pb_subtitles[pb_ind], top_loc[pb_ind]-(title_padding*grid_vert), 
                          bottom_loc[pb_ind], left_loc[pb_ind], right_loc[pb_ind], 15, 12,
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        basic_goal_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in basic_goal_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                basic_goal_df.loc[ind, ca] = basic_goal_lists[iii][ca_ind]
            iii += 1

        ##########     2
        # Generate Viz
        pb_ind = 1
        pb_rg = enrich_df_tasks(allocation_df, pb_tasks)
        curr_c1 = "" + c1
        range_goal_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            range_goal_lists[ind] = list(pb_rg.loc[pb_rg.Task == task,:].Range_val.values)
            range_goal_lists[ind] = [round(i,0) for i in range_goal_lists[ind]]
            curr_c1 += f""".add_yaxis("{task[:5]}", {range_goal_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        range_goal_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in range_goal_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                range_goal_df.loc[ind, ca] = range_goal_lists[iii][ca_ind]
            iii += 1
            
        ##########     3
        # Data Cleaning
        global p_completed_asin
        p_completed_asin = atwt_wp.loc[atwt_wp.PR == 'Proactive-P']
        p_completed_asin = p_completed_asin.groupby(['CA', 'Task']).sum().asins_updated.reset_index()
        for ca in CA_names:
            if ca not in list(p_completed_asin.CA.unique()):
                tmp_task = p_completed_asin.Task.unique()[0]
                tmp_append_ca = pd.DataFrame(data = [[ca, tmp_task, 0]], columns = p_completed_asin.columns)
                p_completed_asin = pd.concat([p_completed_asin, tmp_append_ca], ignore_index=True)
        p_completed_asin = enrich_df_tasks(p_completed_asin, pb_tasks)

        # Generate Viz
        pb_ind = 2
        curr_c1 = "" + c1
        pca_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            pca_lists[ind] = list(p_completed_asin.loc[p_completed_asin.Task == task,:].asins_updated.values)
            print(len(pca_lists[ind]))
            pca_lists[ind] = [round(i,2) for i in pca_lists[ind]]
            curr_c1 += f""".add_yaxis("{task[:5]}", {pca_lists[ind]},color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        pca_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii = 0
        print(len(pca_df), len(CA_names), len(pb_tasks), len(pca_lists))
        for ind, row in pca_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                pca_df.loc[ind, ca] = pca_lists[iii][ca_ind] 
            iii += 1
        
        ##########     4
        # Data Cleaning
        p_touched_new_df = enrich_df_tasks(allo_touched.loc[:,['CA', 'Task', 'ASIN']], pb_tasks) \
                .sort_values(by=['CA', 'Task']) \
                .fillna(value=0)

        # Generate Viz
        pb_ind = 3
        curr_c1 = "" + c1
        pta_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            pta_lists[ind] = list(p_touched_new_df.loc[p_touched_new_df.Task == task,:].ASIN.values)
            pta_lists[ind] = [round(i,2) for i in pta_lists[ind]]
            curr_c1 += f""".add_yaxis("{task[:5]}", {pta_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        pta_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in pta_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                try:
                    pta_df.loc[ind, ca] = pta_lists[iii][ca_ind]
                except:
                    print(ca_ind, ca)
            iii += 1
        pta_df = pta_df.fillna(0)
            
        ##########     5
        # Generate Viz
        pb_ind = 4
        curr_c1 = "" + c1
        pct_bg_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            l, r = pca_lists[ind], basic_goal_lists[ind]
            pct_bg_lists[ind] = [round(l[j]/r[j],2) if r[j]!=0 else 0 for j in range(len(l))]
            
            curr_c1 += f""".add_yaxis("{task[:5]}", {pct_bg_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        pct_bg_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in pct_bg_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                pct_bg_df.loc[ind, ca] = pct_bg_lists[iii][ca_ind]
            iii += 1
        
        ##########     6
        # Generate Viz
        pb_ind = 5
        curr_c1 = "" + c1
        pct_tg_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            l, r = pta_lists[ind], basic_goal_lists[ind]
            pct_tg_lists[ind] = [round(l[j]/r[j],2) if r[j]!=0 else 0 for j in range(len(l))]
            
            curr_c1 += f""".add_yaxis("{task[:5]}", {pct_tg_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c


        ## 数据表格
        pct_tg_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in pct_tg_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                pct_tg_df.loc[ind, ca] = pct_tg_lists[iii][ca_ind]
            iii += 1
        
        ##########     7
        # Data Cleaning
        ra_to_be_df = atwt_wp.loc[atwt_wp.PR == 'Proactive-R',['CA','Task', 'asins_tobe_updated']].reset_index(drop=True)
        for ca in CA_names:
            if ca not in list(ra_to_be_df.CA.unique()):
                tmp_task = ra_to_be_df.Task.unique()[0]
                tmp_append_ca = pd.DataFrame(data = [[ca, tmp_task, 0]], columns = ra_to_be_df.columns)
                ra_to_be_df = pd.concat([ra_to_be_df, tmp_append_ca], ignore_index=True)
        ra_to_be_df = enrich_df_tasks(ra_to_be_df, pb_tasks)

        # Generate Viz
        pb_ind = 6
        curr_c1 = "" + c1
        ra_tobe_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            ra_tobe_lists[ind] = list(ra_to_be_df.loc[ra_to_be_df.Task == task,:].asins_tobe_updated.values)
            curr_c1 += f""".add_yaxis("{task[:5]}", {ra_tobe_lists[ind]},  color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        ra_tobe_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in ra_tobe_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                ra_tobe_df.loc[ind, ca] = ra_tobe_lists[iii][ca_ind]
            iii += 1
        
        ##########     8
        # Data Cleaning
        ra_updated_df = atwt_wp.loc[atwt_wp.PR == 'Proactive-R',['CA','Task', 'asins_updated']].reset_index(drop=True)
        for ca in CA_names:
            if ca not in list(ra_updated_df.CA.unique()):
                tmp_task = ra_updated_df.Task.unique()[0]
                tmp_append_ca = pd.DataFrame(data = [[ca, tmp_task, 0]], columns = ra_updated_df.columns)
                ra_updated_df = pd.concat([ra_updated_df, tmp_append_ca], ignore_index=True)
        ra_updated_df = enrich_df_tasks(ra_updated_df, pb_tasks)

        # Generate Viz
        pb_ind = 7
        curr_c1 = "" + c1
        ra_ed_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            ra_ed_lists[ind] = list(ra_updated_df.loc[ra_updated_df.Task == task,:].asins_updated.values)
            curr_c1 += f""".add_yaxis("{task[:5]}", {ra_ed_lists[ind]},  color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        ra_ed_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in ra_ed_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                ra_ed_df.loc[ind, ca] = ra_ed_lists[iii][ca_ind]
            iii += 1
        
        ##########     9
        # Generate Viz
        pb_ind = 8
        curr_c1 = "" + c1
        pr_ed_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            l, r = pca_lists[ind], ra_ed_lists[ind]
            pr_ed_lists[ind] = [round(l[j] + r[j], 2) for j in range(len(r))]
            
            curr_c1 += f""".add_yaxis("{task[:5]}", {pr_ed_lists[ind]},  color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        pr_ed_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in pr_ed_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                pr_ed_df.loc[ind, ca] = pr_ed_lists[iii][ca_ind]
            iii += 1
        self._signal.emit(96)
        
        ##########     10
        # Generate Viz
        pb_ind = 9
        curr_c1 = "" + c1
        completion_lists = [None for i in pb_tasks]
        for ind, task in enumerate(pb_tasks):
            l, r = pr_ed_lists[ind], basic_goal_lists[ind]
            completion_lists[ind] = [round(l[j]/r[j], 2) if r[j] !=0 else 0 for j in range(len(r))]
            curr_c1 += f""".add_yaxis("{task[:5]}", {completion_lists[ind]}, color="{colors[ind]}", is_selected={default_selected})\n"""

        curr_c2 = c2%(bp_mark_100_line[pb_ind], pb_names[pb_ind], pb_subtitles[pb_ind], 
                      3, 95, 10, 90, 20, 15, 
                      "", 10)
        exec(curr_c1+curr_c2, glb, locs)
        sub_html_dir = fr"{html_dir}\\By_Person__{pb_ind}_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}_{pb_names[pb_ind]}.html"
        c = locs['c']
        c.render(sub_html_dir)

        if pb_ind >= ignore_params:
            hyperlink = f"""title_link=r"{sub_html_dir}", """
            curr_c2 = c2%(bp_mark_100_line[pb_ind],pb_names[pb_ind], pb_subtitles[pb_ind], 
                          top_loc[pb_ind]-(title_padding*grid_vert), bottom_loc[pb_ind], 
                          left_loc[pb_ind], right_loc[pb_ind], 
                          15, 12, 
                          hyperlink, 12)
            exec(curr_c1+curr_c2, glb, locs)
            c = locs['c']
            cs[pb_ind-ignore_params]=c

        ## 数据表格
        completion_df = pd.DataFrame(columns = CA_names, index=pb_tasks)
        iii=0
        for ind, row in completion_df.iterrows():
            for ca_ind, ca in enumerate(CA_names):
                completion_df.loc[ind, ca] = completion_lists[iii][ca_ind]
            iii += 1
        self._signal.emit(99)
        
        # Grid All Sheets
        by_person_grid = Grid(init_opts=opts.InitOpts(width=f"{Grid_width}px",height=f"{Grid_height}px", bg_color = "#FFFFFF"))
        for ind in range(total):
            c = cs[ind]
            by_person_grid.add(
                c, 
                grid_opts=opts.GridOpts(
                    pos_top=f"{top_loc[ind+ignore_params]}%",
                    pos_bottom=f"{bottom_loc[ind+ignore_params]}%",
                    pos_left=f"{left_loc[ind+ignore_params]}%",
                    pos_right=f"{right_loc[ind+ignore_params]}%",
                ), 
                is_control_axis_index=False
            )
        html_name = f"By_Person_{q_wk_from}_{q_wk_to}_{q_year}_{CM_login}.html"
        by_person_grid.render(html_name)
        self.by_person_grid = by_person_grid
        self._name_signal.emit(html_name)
        
        ## Save Data Frame
        by_person_dfs = [
            basic_goal_df, range_goal_df, pca_df, pta_df, pct_bg_df, pct_tg_df,
            ra_tobe_df, ra_ed_df, pr_ed_df, completion_df
        ]
        all_sheets = []
        excel_file_loc = fr"{data_dir}\\By Person REPORT_{q_wk_from}_{q_wk_to}_{q_year}_{the_CM}.xlsx"
        with pd.ExcelWriter(excel_file_loc) as xlsx:
            for task in pb_tasks:
                Person_df = pd.DataFrame()
                for df_ind, df in enumerate(by_person_dfs):
                    df['Metric'] = pb_names[df_ind]
                    df['indexes'] = df_ind
                    curr = pd.DataFrame(df.loc[task,:]).T
                    Person_df = pd.concat([Person_df, curr])
                output = Person_df.groupby(['Metric']).sum().sort_values('indexes').iloc[:, :-1]
                output.to_excel(xlsx, sheet_name=task[:5], index=True)
                all_sheets.append(task[:5])
        revise_workbook(excel_file_loc, CA_names, pb_names, all_sheets)

        ## 字段解释
        # add Explanation
        wb = load_workbook(excel_file_loc)
        if intro_sheet_name not in wb.sheetnames:
            # 创建 Sheet 并塞入内容
            ws = wb.create_sheet(intro_sheet_name, 0)
            ws['A1'] = 'Task Type'
            ws['B1'] = 'Metrics'
            ws['C1'] = 'Intro'
            ws['A2'] = 'Proactive-P'
            ws['A8'] = 'Proactive-R'
            ws['A10'] = 'Summary(P+R)'
            for i in range(len(pb_names)):
                ws[f'B{i+2}'] = pb_names[i]
            for i in range(len(intro_10)):
                ws[f'C{i+2}'] = intro_10[i]

            # 设置行高&列宽
            ws.row_dimensions[1].height=20
            ws.column_dimensions['A'].width=15
            ws.column_dimensions['B'].width=26
            ws.column_dimensions['C'].width=35
            # 字体格式
            #边框
            thin = Side(border_style="thin", color="000000")#边框样式，颜色
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
            #字体
            init_font = Font(size=10, bold=True, name='Microsoft YaHei',  color="101013") #首行首列格式
            cell_font = Font(size=10, bold=False, name='Microsoft YaHei',  color="101013")
            #居中
            cell_align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            for row in ws['A1:C11']:
                for cell in row:
                    if cell.column_letter == 'A' or cell.row == 1:
                        cell.font = init_font
                    else:
                        cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = cell_border
            # 底色
            for letter in ['A', 'B', 'C']:
                ws[f'{letter}1'].fill = fill_1

            # 单元格合并
            ws.merge_cells('A2:A7')
            ws.merge_cells('A8:A9')
            ws.merge_cells('A10:A11')
            wb.save(excel_file_loc)
        print("By Person Report is created!")
        
        self.return_datas = [
            time_compare_grid, by_person_grid
        ]
        self._signal.emit(100)
        qmut_single.unlock()

        
class Thread_multiple_weeks(QThread):
    _signal = pyqtSignal(int)
    _name_signal = pyqtSignal(str) # 用于发送 html 的文件位置
    def __init__(self, datas):
        super().__init__()
        self.datas = datas
        self.return_datas = None
        
    def run(self):
        qmut_multiple.lock()
        q_wk_from, q_wk_to, q_year, \
            weeks, years, \
            the_CM, CM_login, \
            the_CMs, CMs_login, \
            task_list_dir, \
            data_dir, html_dir, \
            multiple_weeks_data, viz_params_by_task, viz_params_overall, \
            tk_names, tk_subtitles, oa_names, oa_subtitles, intro_10 = self.datas
        # 先将数据赋予变量
        overall_data_lists, summarize_weeks_all = multiple_weeks_data
        # INIT
        ignore_params, default_selected, Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, \
            row_sheet_counts, total_metrics, title_padding = viz_params_by_task
        total = total_metrics - ignore_params
        ## Generate Viz
        all_valid_tasks = list(summarize_weeks_all.keys())
        by_task_dfs = []
        
        ## 其余参数通过计算得到
        col_sheet_counts, top_loc, bottom_loc, left_loc, right_loc, Grid_height = \
            get_params(Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, total, ignore_params)
        by_task_grid = Grid(init_opts=opts.InitOpts(width=f"{Grid_width}px",height=f"{Grid_height}px", bg_color = "#FFFFFF"))

        ## 是否添加 100% 的标记线
        hundred_percent_line = """opts.MarkLineItem(name="100%%", y=1, symbol='diamond', symbol_size=[10,10])"""
        mark_100_line = [
            "", "", "", "",
            hundred_percent_line,
            hundred_percent_line,
            "", "", "", 
            hundred_percent_line
        ]
        locs = {}
        glb = {"Bar": Bar, 'opts':opts, 'weeks': weeks}
        c1 = f"""c = (Bar(
                    init_opts=opts.InitOpts(
                        bg_color = "#FFFFFF"
                    )
                )
            .add_xaxis(weeks)\n"""
        c2 = f"""
            .set_series_opts(
                label_opts=opts.LabelOpts(is_show=False),
                markpoint_opts=opts.MarkPointOpts(
                    data=[
                        opts.MarkPointItem(type_="max", name="最大值"),
                    ],
                    symbol="circle",
                    symbol_size=[1,1],
                    label_opts=opts.LabelOpts(
                        position="top", color="#0A0A0D",
                        font_size=9, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

                markline_opts=opts.MarkLineOpts(
                    data=[
                        opts.MarkLineItem(type_="average", name="平均值"),
                        %s
                    ],
                    linestyle_opts = opts.LineStyleOpts(
                        opacity=0.5,
                        width=0.6,
                        type_="dotted",
                    ),
                    symbol_size=[0,1],
                    label_opts=opts.LabelOpts(
                        position="right", 
                        font_size=10, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

            )
            .set_global_opts(
                xaxis_opts=opts.AxisOpts(
                    axislabel_opts=opts.LabelOpts(rotate=0),
                    name='Weeks'
                ),
                yaxis_opts=opts.AxisOpts(
                    name=''
                ),
                title_opts=opts.TitleOpts(title="%s", subtitle="%s", 
                    item_gap=5,
                    pos_top=f"%d%%",
                    pos_bottom=f"%d%%",
                    pos_left=f"%d%%",
                    pos_right=f"%d%%",  
                    title_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    subtitle_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    %s
                ),
                brush_opts=opts.BrushOpts(),
                datazoom_opts=opts.DataZoomOpts(
                    type_="inside",
                    pos_bottom = "10%%",
                    filter_mode="empty"
                ),
                toolbox_opts=opts.ToolboxOpts(
                        pos_left = "40%%",
                        pos_top = "0%%",
                    ),
                legend_opts=opts.LegendOpts(
                        type_ = 'scroll',
                        selected_mode = 'multiple',
                        orient = 'vertical',
                        pos_right = '1%%',
                        pos_left = '{103-right_loc[ignore_params+col_sheet_counts-1]}%%',
                        pos_top = '10%%',
                        align = 'right',
                        item_gap = 20,
                        padding = 5,
                        textstyle_opts = opts.TextStyleOpts(font_size = %d)
                    ),
                tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),

            ))"""
        for tk_ind, tk in enumerate(tqdm(tk_names[:], desc='By Task Report')):
            """
            遍历 10 个 metric，建立 10 个 HTML 和 10 个 DataFrame
            """
            # Generate Viz
            weeks_string_list = [str(i)+'-'+str(j) for i,j in weeks]

            tp=top_loc[tk_ind]
            bot=bottom_loc[tk_ind]
            lft=left_loc[tk_ind]
            rght=right_loc[tk_ind]

            task_cnt = 0
            curr_c1 = "" + c1
            for task, vs in summarize_weeks_all.items():
                curr_c1 += f""".add_yaxis("{task[:5]}", {vs[tk_ind]},  color="{colors[task_cnt]}", is_selected={default_selected})\n"""
                task_cnt += 1
            curr_c2 = c2%(mark_100_line[tk_ind], tk, tk_subtitles[tk_ind], 
                          3, 95, 10, 90, 20, 15, 
                          "", 10)
            exec(curr_c1+curr_c2, glb, locs)
            sub_html_dir = fr"{html_dir}\\By_Task__{tk_ind}_{tk}.html"
            c = locs['c']
            c.render(sub_html_dir)

            if tk_ind >= ignore_params:
                hyperlink = f"""title_link=r"{sub_html_dir}", """
                curr_c2 = c2%(mark_100_line[tk_ind], 
                              tk, tk_subtitles[tk_ind], 
                              tp-(title_padding*grid_vert), bot, lft, rght,
                              15, 12, 
                              hyperlink, 12)
                exec(curr_c1+curr_c2, glb, locs)
                c = locs['c']
                by_task_grid.add(
                    c, 
                    grid_opts=opts.GridOpts(
                        pos_top=f"{tp}%",
                        pos_bottom=f"{bot}%",
                        pos_left=f"{lft}%",
                        pos_right=f"{rght}%",
                    ), 
                    is_control_axis_index=False
                )

            ## 数据表格
            tk_df = pd.DataFrame(columns=weeks_string_list, index=all_valid_tasks)
            iii=0
            for task, row in tk_df.iterrows():
                for week_str_ind, week_str in enumerate(weeks_string_list):
                    tk_df.loc[task, week_str] = summarize_weeks_all[task][tk_ind][week_str_ind]
                iii += 1
            by_task_dfs.append(tk_df)
            self._signal.emit(10 + tk_ind*4)
        html_name = f"By_Task_{years[0]}_{weeks[0][0]}_{years[-1]}_{weeks[-1][-1]}.html"
        by_task_grid.render(html_name)
        self._name_signal.emit(html_name)
        
        # By Task Data Frame
        all_sheets = []
        excel_file_loc = fr"{data_dir}\\By Task REPORT_{years[0]}_{weeks[0][0]}_{years[-1]}_{weeks[-1][-1]}.xlsx"
        with pd.ExcelWriter(excel_file_loc) as xlsx:
            for task in all_valid_tasks:
                Task_df = pd.DataFrame()
                for df_ind, df in enumerate(by_task_dfs):
                    df['Metric'] = tk_names[df_ind]
                    df['indexes'] = df_ind
                    curr = pd.DataFrame(df.loc[task,:]).T
                    Task_df = pd.concat([Task_df, curr])
                output = Task_df.groupby(['Metric']).sum().sort_values('indexes').iloc[:, :-1]
                output.to_excel(xlsx, sheet_name=task[:5], index=True)
                all_sheets.append(task[:5])
        revise_workbook(excel_file_loc, weeks, tk_names, all_sheets)    
        ## 字段解释
        wb = load_workbook(excel_file_loc)
        intro_sheet_name = 'Intro'
        if intro_sheet_name not in wb.sheetnames:
            # 创建 Sheet 并塞入内容
            ws = wb.create_sheet(intro_sheet_name, 0)
            ws['A1'] = 'Task Type'
            ws['B1'] = 'Metrics'
            ws['C1'] = 'Intro'
            ws['A2'] = 'Proactive-P'
            ws['A8'] = 'Proactive-R'
            ws['A10'] = 'Summary(P+R)'
            for i in range(len(tk_names)):
                ws[f'B{i+2}'] = tk_names[i]
            for i in range(len(intro_10)):
                ws[f'C{i+2}'] = intro_10[i]

            # 设置行高&列宽
            ws.row_dimensions[1].height=20
            ws.column_dimensions['A'].width=15
            ws.column_dimensions['B'].width=26
            ws.column_dimensions['C'].width=35
            # 字体格式
            #边框
            thin = Side(border_style="thin", color="000000")#边框样式，颜色
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
            #字体
            init_font = Font(size=10, bold=True, name='Microsoft YaHei',  color="101013") #首行首列格式
            cell_font = Font(size=10, bold=False, name='Microsoft YaHei',  color="101013")
            #居中
            cell_align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            for row in ws['A1:C11']:
                for cell in row:
                    if cell.column_letter == 'A' or cell.row == 1:
                        cell.font = init_font
                    else:
                        cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = cell_border
            # 底色
            for letter in ['A', 'B', 'C']:
                ws[f'{letter}1'].fill = fill_1

            # 单元格合并
            ws.merge_cells('A2:A7')
            ws.merge_cells('A8:A9')
            ws.merge_cells('A10:A11')
            wb.save(excel_file_loc)
        print("By Task Report is created!")
        # INIT
        ignore_params, default_selected, Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, \
            row_sheet_counts, total_metrics, title_padding = viz_params_overall
        total = total_metrics - ignore_params
        ## Overall, CMs
        all_valid_titles = overall_data_lists.keys()
        weeks_string_list = [str(i)+'-'+str(j) for i,j in weeks]
        overall_dfs = []
        ## 其余参数通过计算得到
        col_sheet_counts, top_loc, bottom_loc, left_loc, right_loc, Grid_height = \
            get_params(Grid_width, padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, total, ignore_params)
        overall_grid = Grid(init_opts=opts.InitOpts(width=f"{Grid_width}px",height=f"{Grid_height}px", bg_color = "#FFFFFF"))
        c1 = f"""c = (Bar(
                    init_opts=opts.InitOpts(
                        bg_color = "#FFFFFF"
                    )
                )
            .add_xaxis(weeks)\n"""
        c2 = f"""
            .set_series_opts(
                label_opts=opts.LabelOpts(is_show=False),
                markpoint_opts=opts.MarkPointOpts(
                    data=[
                        opts.MarkPointItem(type_="max", name="最大值"),
                    ],
                    symbol="circle",
                    symbol_size=[1,1],
                    label_opts=opts.LabelOpts(
                        position="top", color="#0A0A0D",
                        font_size=9, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

                markline_opts=opts.MarkLineOpts(
                    data=[
                        opts.MarkLineItem(type_="average", name="平均值"),
                        %s
                    ],
                    linestyle_opts = opts.LineStyleOpts(
                        opacity=0.5,
                        width=0.6,
                        type_="dotted",
                    ),
                    symbol_size=[0,1],
                    label_opts=opts.LabelOpts(
                        position="right", 
                        font_size=10, font_style='normal', font_weight='bold',
                        font_family="Microsoft YaHei",margin=8
                    )
                ),

            )
            .set_global_opts(
                xaxis_opts=opts.AxisOpts(
                    axislabel_opts=opts.LabelOpts(rotate=0),
                    name='Weeks'
                ),
                yaxis_opts=opts.AxisOpts(
                    name=''
                ),
                title_opts=opts.TitleOpts(title="%s", subtitle="%s", 
                    item_gap=5,
                    pos_top=f"%d%%",
                    pos_bottom=f"%d%%",
                    pos_left=f"%d%%",
                    pos_right=f"%d%%",  
                    title_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    subtitle_textstyle_opts = opts.TextStyleOpts(font_size = %d),
                    %s
                ),
                brush_opts=opts.BrushOpts(),
                datazoom_opts=opts.DataZoomOpts(
                    type_="inside",
                    pos_bottom = "10%%",
                    filter_mode="empty"
                ),
                toolbox_opts=opts.ToolboxOpts(
                        pos_left = "40%%",
                        pos_top = "0%%",
                    ),
                legend_opts=opts.LegendOpts(
                        type_ = 'scroll',
                        selected_mode = 'multiple',
                        orient = 'vertical',
                        pos_right = '1%%',
                        pos_left = '{103-right_loc[ignore_params+col_sheet_counts-1]}%%',
                        pos_top = '10%%',
                        align = 'right',
                        item_gap = 20,
                        padding = 5,
                        textstyle_opts = opts.TextStyleOpts(font_size = %d)
                    ),
                tooltip_opts=opts.TooltipOpts(trigger="axis", axis_pointer_type="cross"),

            ))"""

        for oa_ind, oa in enumerate(tqdm(oa_names[:], desc='Overall Report')):
            tp=top_loc[oa_ind]
            bot=bottom_loc[oa_ind]
            lft=left_loc[oa_ind]
            rght=right_loc[oa_ind]

            curr_c1 = "" + c1
            color_cnt = 0
            for x_ind, x in enumerate(all_valid_titles):
                vals = overall_data_lists[x][oa_ind]
                vals = [round(i,2) for i in vals]
                curr_c1 += f""".add_yaxis("{x}", {vals},  color="{colors[color_cnt]}", is_selected={default_selected})\n"""
                color_cnt += 1

            curr_c2 = c2%(mark_100_line[oa_ind], oa, oa_subtitles[oa_ind], 3, 95, 10, 90, 20, 15, "", 10)
            exec(curr_c1+curr_c2, glb, locs)
            sub_html_dir = fr"{html_dir}\\Overall__{oa_ind}_{oa}.html"
            c = locs['c']
            c.render(sub_html_dir)

            if oa_ind >= ignore_params:
                hyperlink = f"""title_link=r"{sub_html_dir}", """
                curr_c2 = c2%(mark_100_line[oa_ind], oa, oa_subtitles[oa_ind], tp-(title_padding*grid_vert), 
                              bot, lft, rght, 15, 12, hyperlink, 12)
                exec(curr_c1+curr_c2, glb, locs)
                c = locs['c']
                overall_grid.add(
                    c, 
                    grid_opts=opts.GridOpts(
                        pos_top=f"{tp}%",
                        pos_bottom=f"{bot}%",
                        pos_left=f"{lft}%",
                        pos_right=f"{rght}%",
                    ), 
                    is_control_axis_index=False
                )

            ## 数据表格
            oa_df = pd.DataFrame(columns=weeks_string_list, index=all_valid_titles)
            iii=0
            for title, row in oa_df.iterrows():
                for week_str_ind, week_str in enumerate(weeks_string_list):
                    oa_df.loc[title, week_str] = overall_data_lists[title][oa_ind][week_str_ind]
                iii += 1
            overall_dfs.append(oa_df)
            self._signal.emit(55 + oa_ind*3.6)
        html_name = f"Overall_{years[0]}_{weeks[0][0]}_{years[-1]}_{weeks[-1][-1]}.html"
        overall_grid.render(html_name)
        self._name_signal.emit(html_name)
        
        # Overall Data Frame
        all_sheets = []
        excel_file_loc = fr"{data_dir}\\Overall REPORT_{years[0]}_{weeks[0][0]}_{years[-1]}_{weeks[-1][-1]}.xlsx"
        with pd.ExcelWriter(excel_file_loc) as xlsx:
            for title in all_valid_titles:
                CM_df = pd.DataFrame()
                for df_ind, df in enumerate(overall_dfs):
                    df['Metric'] = oa_names[df_ind]
                    df['indexes'] = df_ind
                    curr = pd.DataFrame(df.loc[title,:]).T
                    CM_df = pd.concat([CM_df, curr])
                output = CM_df.groupby(['Metric']).sum().sort_values('indexes').iloc[:, :-1]
                output.to_excel(xlsx, sheet_name=title, index=True)
                all_sheets.append(title)        
        revise_workbook(excel_file_loc, weeks, oa_names, all_sheets)    
        ## 字段解释
        wb = load_workbook(excel_file_loc)
        if intro_sheet_name not in wb.sheetnames:
            # 创建 Sheet 并塞入内容
            ws = wb.create_sheet(intro_sheet_name, 0)
            ws['A1'] = 'Task Type'
            ws['B1'] = 'Metrics'
            ws['C1'] = 'Intro'
            ws['A2'] = 'Proactive-P'
            ws['A8'] = 'Proactive-R'
            ws['A10'] = 'Summary(P+R)'
            for i in range(len(oa_names)):
                ws[f'B{i+2}'] = oa_names[i]
            for i in range(len(intro_10)):
                ws[f'C{i+2}'] = intro_10[i]

            # 设置行高&列宽
            ws.row_dimensions[1].height=20
            ws.column_dimensions['A'].width=15
            ws.column_dimensions['B'].width=26
            ws.column_dimensions['C'].width=35
            # 字体格式
            #边框
            thin = Side(border_style="thin", color="000000")#边框样式，颜色
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)#边框的位置
            #字体
            init_font = Font(size=10, bold=True, name='Microsoft YaHei',  color="101013") #首行首列格式
            cell_font = Font(size=10, bold=False, name='Microsoft YaHei',  color="101013")
            #居中
            cell_align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            for row in ws['A1:C11']:
                for cell in row:
                    if cell.column_letter == 'A' or cell.row == 1:
                        cell.font = init_font
                    else:
                        cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = cell_border
            # 底色
            for letter in ['A', 'B', 'C']:
                ws[f'{letter}1'].fill = fill_1

            # 单元格合并
            ws.merge_cells('A2:A7')
            ws.merge_cells('A8:A9')
            ws.merge_cells('A10:A11')
            wb.save(excel_file_loc)
        print("Overall Report is created!")
        
        self.return_datas = [by_task_grid, overall_grid]
        
        self._signal.emit(100)
        qmut_multiple.unlock()
        
        
class Thread_Integrate(QThread):
    _signal =pyqtSignal(int)
    def __init__(self, datas):
        super().__init__()
        self.datas = datas
        
    def run(self):
        self._signal.emit(0)
        tab1, tab2, tab3, tab4 = self.datas[:4]
        title0, title1, title2, title3, title4, file_title = self.datas[4]
        qmut_integrate.lock()
        # Gather ALL
        tab = Tab(page_title=title0)
        tab.add(tab1, title1)
        self._signal.emit(30)
        tab.add(tab2, title2)
        self._signal.emit(50)
        tab.add(tab3, title3)
        self._signal.emit(80)
        tab.add(tab4, title4)
        self._signal.emit(90)
        tab.render(file_title)
        self._signal.emit(100)
        print('The Integrated Report is created!')
        qmut_integrate.unlock()
        
        
class MYGUI(QTabWidget):
    def __init__(self):
        super().__init__()
        self.exit_flag = False
        self.tabUI()
        self.test_counter = 0

        self.root = None
        self.weeks = []
        self.years = []
        self.group_cnt = 0
        self.single_week_data, self.multiple_weeks_data = None, None
        
        # 五份报告预设为 None
        self.htmls_attachments = []
#         self.time_compare_grid = None
#         self.by_person_grid = None
#         self.by_task_grid = None
#         self.overall_grid = None
#         self.integrated_html_grid = None
        
    def tabUI(self):
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tab3 = QWidget()
        
        self.addTab(self.tab1,'Upload Data')
        self.addTab(self.tab2,'Download Reports')
        self.addTab(self.tab3,'Email Reports')
        
        self.upload_UI()
        self.report_UI()
        self.emails_UI()
        
    def upload_UI(self):
        """
        设置两个垂直的Layout？
        左侧放 QLabel，右侧放 QLineEdit
        """
        self.Inputs = []
        
        # 设置 3 个输入框， 1 个勾选框
        self.wanted_task_label = QLabel("Task前缀(必填)(查询用)")
        self.wanted_task_label.setFont(label_font)
        self.wanted_task_label.setToolTip("如: JP123")
        self.wanted_task_path = QLineEdit("JP")
        self.wanted_task_path.setEnabled(True)
        self.wanted_task_path.setMaxLength(6)
        self.wanted_task_path.setPlaceholderText("e.g. JP123")
        self.wanted_task_path.setToolTip("查询条件: Task")
        
        self.wanted_week_label = QLabel("Week数(必填)")
        self.wanted_week_label.setFont(label_font)
        self.wanted_week_label.setToolTip("若全选则输入'A', 否则请输入一个或更多的数字，用逗号(,)隔开")
        self.wanted_week_path = QLineEdit("")
        self.wanted_week_path.setEnabled(True)
        self.wanted_week_path.setMaxLength(20)
        self.wanted_week_path.setPlaceholderText("e.g. 1,2,3")
        
        self.wanted_file_label = QLabel("Excel 文件名命名规则(可填)")
        self.wanted_file_label.setFont(label_font)
        self.wanted_file_label.setToolTip("输入'_2021'筛选文件名中含有此字段的文件，若有多组规则，用逗号(,)隔开")
        self.wanted_file_path = QLineEdit("")
        self.wanted_file_path.setEnabled(True)
        self.wanted_file_path.setMaxLength(55)
        self.wanted_file_path.setPlaceholderText("e.g. _2021,_2022,_2023")
        
        self.wanted_check = QCheckBox("添加数据识别查重环节(全新数据不需要)")

        self.year_label = QLabel("Year(查询用)")
        self.year_label.setFont(label_font)
        self.year_path = QLineEdit(f"{moment_year}")
        self.year_path.setEnabled(True)
        self.year_path.setMaxLength(4)
        self.year_path.setToolTip(f"查询条件: 年份")
        
        self.refresh_connect_btn = QPushButton("数据库重连")
        self.refresh_connect_btn.setFont(label_font)
        self.refresh_connect_btn.setStyleSheet(run_button_Style)
        icon_dir = r"C:\Users\zihaoz\Documents\PythonScripts\zihaozPython\AVS_proactive\reconnect_btn.png"
        self.refresh_connect_btn.setIcon(QIcon(icon_dir))
        self.refresh_connect_btn.setToolTip("Re-Connect to Database")
        self.refresh_connect_btn.clicked.connect(self.ReconnectDataBase)
        
        self.scout_btn = QPushButton("  查询  ")
        self.scout_btn.setStyleSheet(run_button_Style)
        self.scout_btn.setToolTip("点击查询已上传的工作周，结果在下方显示")
        self.scout_btn.clicked.connect(self.ScoutWeeks)
        
        # QPushButton 放在一个 QHBoxLayout 中
        ## 1. Add
        self.btn_h = QHBoxLayout()
        self.input_btn = QPushButton("  Add  ")
        self.input_btn.setStyleSheet(run_button_Style)
        self.input_btn.clicked.connect(self.AddInput)
        self.input_btn.setToolTip("点击此按钮将数据写入右侧框中！")
        self.input_btn.setStatusTip("状态按钮？")
        ## 2. Upload
        self.upload_btn = QPushButton("  Upload Data  ")
        self.upload_btn.setStyleSheet(run_button_Style)
        self.upload_btn.clicked.connect(self.Upload_Touched_thread)
        self.upload_btn.setToolTip("点击此按钮上传 Touched 数据")
        self.btn_h.addWidget(self.input_btn)
        self.btn_h.addWidget(self.upload_btn)
        
        # 数据读取自此
        self.input_info = QTextEdit("")
        self.input_info.setReadOnly(False)
        self.input_info.setPlaceholderText("上传数据参数队列")
        self.input_info.setToolTip("上传数据参数队列")
        
        # 上传 CM data
        self.cm_btn = QPushButton("  Upload CM Data  ")
        self.cm_btn.setStyleSheet(run_button_Style)
        self.cm_btn.clicked.connect(self.Upload_CM_thread)
        self.cm_btn.setToolTip("上传 CM Data")
        
        # Bug 提示框
        self.bug_label = QLabel("Log")
        self.bug_label.setFont(label_font)
        self.bug_info = QTextEdit("")
        self.bug_info.setReadOnly(True)
        self.bug_info.setPlaceholderText("Log Information...")

        # Progress 提示框
        self.progress_label = QLabel("Progress")
        self.progress_label.setFont(label_font)
        self.progress_info = QTextEdit("")
        self.progress_info.setReadOnly(True)
        self.progress_info.setPlaceholderText("Progress Information...")
        
        # DIY Progress Bar
        self.t1_pbar = QLineEdit("")
        self.t1_pbar.setMaxLength(10)
        self.t1_pbar.setEnabled(True)
        self.t1_pbar.setToolTip(f"Show the progress")
        
        # 0.总布局
        vs = QVBoxLayout()
        # 1.将输入框放入 垂直布局中
        hs = QHBoxLayout()
        v1 = QVBoxLayout()
        v2 = QVBoxLayout()
        ## Task & Week & File & Check
        v1.addWidget(self.wanted_task_label)
        v1.addWidget(self.wanted_week_label)
        v1.addWidget(self.wanted_file_label)
        v1.addWidget(self.wanted_check, 0, Qt.AlignTop | Qt.AlignRight)
        v1.addStretch(1)
        v1.addWidget(self.cm_btn)
        
        # Year 查询
        h_scout_year = QHBoxLayout()
        h_scout_year.addWidget(self.year_path)
        h_scout_year.addWidget(self.year_label)
        # 重连数据库 & Scout
        h_scout = QHBoxLayout()
        h_scout.addWidget(self.refresh_connect_btn)
        h_scout.addWidget(self.scout_btn)
        
        v2.addWidget(self.wanted_task_path)
        v2.addWidget(self.wanted_week_path)
        v2.addWidget(self.wanted_file_path)
        v2.addLayout(self.btn_h)
        v2.addLayout(h_scout_year)
        v2.addLayout(h_scout)
        
        hs.addStretch(1)
        hs.addLayout(v1)
        hs.addStretch(4)
        hs.addLayout(v2)
        hs.addStretch(4)
        hs.addWidget(self.input_info)
        hs.addStretch(1)
#         hs.setSpacing(30)
        
        # 2.文本框
        ## Label
        h_text_layout = QHBoxLayout()
        v_label_layout = QVBoxLayout()
        v_label_layout.addWidget(self.bug_label, 0, Qt.AlignCenter)
        v_label_layout.addWidget(self.progress_label, 0, Qt.AlignCenter)
        ## Info Text
        v_info_layout = QVBoxLayout()
        v_info_layout.addWidget(self.bug_info)
        v_info_layout.addWidget(self.progress_info)
        
        h_text_layout.addLayout(v_label_layout)
        h_text_layout.addLayout(v_info_layout)
        
        # 3.总布局收纳所有框框
        vs.addStretch(1)
        vs.addLayout(hs)
        vs.addStretch(1)
        vs.addLayout(h_text_layout)   
        vs.addWidget(self.t1_pbar)
        vs.addStretch(5)
        # 4.设定总布局
        self.tab1.setLayout(vs)
        
    def report_UI(self):
        default_params_file = r"\\ant\dept-as\PEK10\DEPT2\RBS\AVS\AVS_Task Related\Proactive Task\[Updating] Proactive work allocation\Report_GUI\Report_Params\Params.xlsx"
        try:
            params_df = pd.read_excel(default_params_file,
                                sheet_name='Dirs', index_col='Index')
            task_list_dir = params_df.loc['Task_list_dir', 'URL']
            CM_info_default = params_df.loc['CM_info_dir', 'URL']
            root_path = params_df.loc['Export_dir', 'URL'].strip()
            if not root_path or str(root_path) == 'nan' :
                root_path = os.getcwd()
        except:
            print(default_params_file)
            print("Cannot locate the dafault params excel file!")
            root_path = os.getcwd()
            task_list_dir = r"\\ant\dept-as\PEK10\DEPT2\RBS\AVS\SVS\Team Members\XingYing\Proactive Completion Report Template.xlsx"
            CM_info_default = r"\\ant\dept-as\PEK10\DEPT2\RBS\AVS\AVS_Task Related\Proactive Task\[Updating] Proactive work allocation\Work Allocation File\CM_name_login.xlsx"

        # 报告导出路径 文件夹
        self.root_root_path = root_path
        select_label_txt = "  Select  "
        self.root_label = QLabel("报告导出路径")
        self.root_label.setFont(label_font)
        self.root_label.setToolTip("根文件夹路径")
        self.root_btn = QPushButton(select_label_txt)
        self.root_btn.setToolTip("Select Folder")
        self.root_btn.clicked.connect(self.choose_root_dir)
        self.root_btn.setStyleSheet(open_folder_button_Style)
        self.root_path = QLineEdit(root_path)
        self.root_path.setCursorPosition(0)
        self.root_path.setEnabled(True)
        # Task List Excel 文件
        self.input_label = QLabel("Task List Excel File")
        self.input_label.setFont(label_font)
        self.input_label.setToolTip("默认存在于 XingYing 的公盘中，若有变更请及时沟通")
        self.input_btn = QPushButton(select_label_txt)
        self.input_btn.setToolTip("Select File")
        self.input_btn.clicked.connect(self.open_source_file)
        self.input_btn.setStyleSheet(open_folder_button_Style)
        self.input_path = QLineEdit(task_list_dir)
        self.input_path.setCursorPosition(0)
        self.input_path.setEnabled(True)
        # CM info Excel 文件
        self.CM_label = QLabel("CM Info Excel File")
        self.CM_label.setFont(label_font)
        self.CM_label.setToolTip("The data of all CMs will be extracted for reports of multiple weeks")
        self.CM_btn = QPushButton(select_label_txt)
        self.CM_btn.setToolTip("Select File")
        self.CM_btn.clicked.connect(self.open_CM_info_file)
        self.CM_btn.setStyleSheet(open_folder_button_Style)
        self.CM_path = QLineEdit(CM_info_default)
        self.CM_path.setCursorPosition(0)
        self.CM_path.setEnabled(True)
        # 参数 Excel 文件
        self.params_file_label = QLabel('Params Excel File')
        self.params_file_label.setFont(label_font)
        self.params_file_label.setToolTip("Revise the parameters and DIY your report")
        self.params_file_btn = QPushButton(select_label_txt)
        self.params_file_btn.setToolTip("Select File")
        self.params_file_btn.clicked.connect(self.open_params_file)
        self.params_file_btn.setStyleSheet(open_folder_button_Style)
        self.params_file_path = QLineEdit(default_params_file)
        self.params_file_path.setCursorPosition(0)
        self.params_file_path.setEnabled(True)

        # 分割线
        # 'en', 'es', 'pt', 'it'
        random_emojis = []
        for k,v in my_emoji['it'].items():
            seed = random.randint(1,100)
            if seed > 50:
                random_emojis.append(v)    
            if len(random_emojis) == 200:
                break
        random.shuffle(random_emojis)
        split_line = " ".join(random_emojis)
        self.split_label = QLabel(split_line)
        self.split_label.setToolTip("a line of adorable emoji~")
        
        # 单周期报告标题
        self.single_week_seperate_label = QLabel("By Person & Time Compare Report(单周期)")
        self.single_week_seperate_label.setFont(label_font)
        self.single_week_seperate_label.setToolTip("单周期报告 (1 CM)")
        
        # 预设的 CM 和 日期信息
        premises = pd.read_excel(default_params_file, sheet_name='CM&Time')
        self.the_CM_label = QLabel("The CM")
        self.the_CM_path = QLineEdit("")
        self.the_CM_path.setEnabled(True)
        self.the_CM_path.setMaxLength(10)
        self.the_CM_path.setPlaceholderText(premises.loc[0, 'The_CM'])
        
        self.CM_login_label = QLabel("CM Login")
        self.CM_login_path = QLineEdit("")
        self.CM_login_path.setEnabled(True)
        self.CM_login_path.setMaxLength(10)
        self.CM_login_path.setPlaceholderText(premises.loc[0, 'CM_login'])
        
        self.week_from_label = QLabel("Week From")
        self.week_from_path = QLineEdit("")
        self.week_from_path.setPlaceholderText("1")
        self.week_from_path.setEnabled(True)
        self.week_from_path.setMaxLength(2)
        
        self.week_to_label = QLabel("Week To")
        self.week_to_path = QLineEdit("")
        self.week_to_path.setPlaceholderText("53")
        self.week_to_path.setEnabled(True)
        self.week_to_path.setMaxLength(2)
        
        self.year_label = QLabel("Year")
        self.year_path = QLineEdit("")
        self.year_path.setPlaceholderText(f"{moment_year}")
        self.year_path.setEnabled(True)
        self.year_path.setMaxLength(4)
        
        # 多周期报告标题
        self.multiple_weeks_seperate_label = QLabel("Overall & By Task Report(多周期)")
        self.multiple_weeks_seperate_label.setFont(label_font)
        self.multiple_weeks_seperate_label.setToolTip("多周期报告（All CMs）")
        
        self.multiple_weeks_input = QTextEdit("")
        self.multiple_weeks_input.setPlaceholderText("输入参考格式: Week_from + 空格 + Week_to + 空格 + Year")
        self.multiple_weeks_input.setReadOnly(False)
        
        self.run_btn = QPushButton("  1. Get Data  ")
        self.run_btn.setToolTip("点击获取数据，该步骤相对耗时")
        self.run_btn.setStyleSheet(run_button_Style)
        self.run_btn.clicked.connect(self.run)
        self.run_btn.setCheckable(True)
        
        self.single_week_report_btn = QPushButton("  2. Generate By Person / Time Compare Report  ")
        self.single_week_report_btn.setToolTip("点击制作单周期报告")
        self.single_week_report_btn.setStyleSheet(run_button_Style)
        self.single_week_report_btn.clicked.connect(self.run_single_week_report)
        
        self.multiple_weeks_report_btn = QPushButton("  3. Generate Overall / By Task Report  ")
        self.multiple_weeks_report_btn.setToolTip("点击制作多周期报告")
        self.multiple_weeks_report_btn.setStyleSheet(run_button_Style)
        self.multiple_weeks_report_btn.clicked.connect(self.run_multiple_weeks_report)
        
        self.all_report_html_btn = QPushButton("  4. Integrate Reports  ")
        self.all_report_html_btn.setToolTip("点击整合所有报告内容")
        self.all_report_html_btn.setStyleSheet(run_button_Style)
        self.all_report_html_btn.clicked.connect(self.all_report_html)
        
        self.info_label = QLabel("Description")
        self.info_label.setFont(label_font)
        self.info_label.setToolTip(info_tool_tip)
        
        # 信息展示
        self.info = QPlainTextEdit()
        self.info.appendPlainText(info_tool_tip)
#         self.info.appendHtml(info_html)
        self.info.setReadOnly(True)
        self.info_cursor = self.info.textCursor()
        self.info_cursor.movePosition(QTextCursor.Start)
        self.info.setTextCursor(self.info_cursor)
        self.info.ensureCursorVisible()
        
        # 进度条
        self.pbar = QProgressBar(self)
        self.pbar.setValue(0)
        self.pbar.setToolTip("一根不中看的进度条")
        # 退出按钮
        exit_btn = QPushButton("  Exit  ")
        exit_btn.setStyleSheet(exit_button_Style)
        exit_btn.clicked.connect(self.Exit)    
        exit_btn.setToolTip("点我快速下班")
        
        
        # 文件 & 文件夹 Input 的 QHBoxLayout
        h_file = QHBoxLayout() # 水平排列布局
        v_label = QVBoxLayout()
        v_path = QVBoxLayout()
        v_btn = QVBoxLayout()
        
        v_label.addWidget(self.root_label)
        v_path.addWidget(self.root_path)
        v_btn.addWidget(self.root_btn)
               
        v_label.addWidget(self.input_label)
        v_path.addWidget(self.input_path)
        v_btn.addWidget(self.input_btn)

        v_label.addWidget(self.CM_label)
        v_path.addWidget(self.CM_path)
        v_btn.addWidget(self.CM_btn)
        
        v_label.addWidget(self.params_file_label)
        v_path.addWidget(self.params_file_path)
        v_btn.addWidget(self.params_file_btn)
        
        h_file.addLayout(v_label)
        h_file.addLayout(v_path)
        h_file.addLayout(v_btn)

        # 单周 数据分割线
        h_single_week_title = QHBoxLayout() # 水平排列布局
        h_single_week_title.addWidget(self.single_week_seperate_label)
        
        # 单周 CM Info
        h_single_week = QHBoxLayout()
        v_t1 = QVBoxLayout()
        v_t2 = QVBoxLayout()
        v_t3 = QVBoxLayout()
        v_l1 = QVBoxLayout()
        v_l2 = QVBoxLayout()
        v_l3 = QVBoxLayout()
        
        v_t1.addWidget(self.the_CM_label)
        v_l1.addWidget(self.the_CM_path)
        v_t1.addWidget(self.CM_login_label)
        v_l1.addWidget(self.CM_login_path)

        v_t2.addWidget(self.week_from_label)
        v_l2.addWidget(self.week_from_path)
        v_t2.addWidget(self.week_to_label)
        v_l2.addWidget(self.week_to_path)
        v_t3.addWidget(self.year_label, 0, Qt.AlignBottom)
        v_l3.addWidget(self.year_path, 0, Qt.AlignBottom)
        
        h_single_week.addStretch(1)
        h_single_week.addLayout(v_t1)
        h_single_week.addLayout(v_l1)
        h_single_week.addStretch(3)
        h_single_week.addLayout(v_t2)
        h_single_week.addLayout(v_l2)
        h_single_week.addStretch(3)
        h_single_week.addLayout(v_t3)
        h_single_week.addLayout(v_l3)
        h_single_week.addStretch(1)
    
        # 多周 数据分割线
        h_multiple_weeks_title = QHBoxLayout() # 水平排列布局
        h_multiple_weeks_title.addWidget(self.multiple_weeks_seperate_label)
        
        # 多周 数据输入
        h_multiple_weeks = QHBoxLayout() # 水平排列布局
        h_multiple_weeks.addWidget(self.multiple_weeks_input)
        
        # 运行
        h9 = QHBoxLayout()
        h9.addWidget(self.run_btn)
        h9.addStretch(1)
        h9.addWidget(self.single_week_report_btn)
        h9.addStretch(1)
        h9.addWidget(self.multiple_weeks_report_btn)
        h9.addStretch(1)
        h9.addWidget(self.all_report_html_btn)
        
        # Info Label
        h10 = QHBoxLayout()
        h10.addWidget(self.info_label)
        
        h100 = QHBoxLayout()
        h100.addStretch(1)
        h100.addWidget(exit_btn)
        
        v = QVBoxLayout() # 垂直排列布局
        v.addLayout(h_file)
        v.addWidget(self.split_label, 0, Qt.AlignCenter)
        v.addLayout(h_single_week_title)
        v.addLayout(h_single_week)
        v.addLayout(h_multiple_weeks_title)
        v.addLayout(h_multiple_weeks)
        v.addLayout(h9)
        v.addLayout(h10)
        v.addWidget(self.info)
        v.addWidget(self.pbar)
        v.addLayout(h100)
        
        self.tab2.setLayout(v)
        screen_w = int(QDesktopWidget().screenGeometry().width())
        screen_h = int(QDesktopWidget().screenGeometry().height())
        width = screen_w/1.7
        height = screen_h/1.2
        x = (screen_w-width)/2
        y = (screen_h-height)/2
        self.setGeometry(x, y, width, height)
        self.setWindowTitle('History File Report')
        
        ## 当前是第几周
        curr = datetime.datetime.now().strftime("%Y/%m/%d")
        curr_week = int(time.strftime("%W").strip()) # 今年没有第一周
        curr_year = int(time.strftime("%Y").strip())
        self.curr_week, self.curr_year = curr_week, curr_year
        self.t2_text(f'{curr} --> This is the {curr_week}th Week')   
        
        self.show()

    def emails_UI(self): 
        self.receivers_label = QLabel("收件人")
        self.receivers_label.setFont(label_font)
        self.receivers_label.setToolTip("Receivers")
        self.receivers_input = QLineEdit("")
        self.receivers_input.setEnabled(True)
        self.receivers_input.setPlaceholderText("请输入收件人邮件地址，并以分号间隔")
        
        self.CCs_label = QLabel("CCs")
        self.CCs_label.setFont(label_font)
        self.CCs_label.setToolTip("Carbon copy")
        self.CCs_input = QLineEdit("")
        self.CCs_input.setEnabled(True)
        self.CCs_input.setPlaceholderText("请输入抄送人邮件地址，并以分号间隔")
        
        self.subject_label = QLabel("邮件主题")
        self.subject_label.setFont(label_font)
        self.subject_label.setToolTip("Topic of this email")
        self.subject_input = QLineEdit("")
        self.subject_input.setEnabled(True)
        self.subject_input.setPlaceholderText("请输入邮件主题")
        
        self.email_title = QLabel(" 邮件正文 ")
        self.email_title.setFont(label_font)
        self.email_title.setToolTip("The content of your email")
        self.emial_text_input = QTextEdit("")
        self.emial_text_input.setEnabled(True)
        self.emial_text_input.setPlaceholderText("输入邮件正文")
        
        self.send_btn = QPushButton(" Draft ")
        self.send_btn.clicked.connect(self.EmailSender)
        self.send_btn.setStyleSheet(run_button_Style)
        self.send_btn.setToolTip("点击将自动起草邮件，请确认无误后发出！")
        
        self.email_info = QTextEdit("")
        self.email_info.setReadOnly(True)
        self.email_info.setPlaceholderText("Progress Information...")
        self.email_info.setToolTip("What to expect here?")
        
        # 0.总布局
        vs = QVBoxLayout()
        
        # 1. 收件人、抄送人
        v_people = QHBoxLayout()
        
        v_rec = QVBoxLayout()
        v_ccs = QVBoxLayout()
        v_sub = QVBoxLayout()
        
        v_rec.addWidget(self.receivers_label, 0, Qt.AlignTop | Qt.AlignLeft)
        v_rec.addWidget(self.receivers_input)
        
        v_ccs.addWidget(self.CCs_label, 0, Qt.AlignTop | Qt.AlignLeft)
        v_ccs.addWidget(self.CCs_input)
        
        v_sub.addWidget(self.subject_label, 0, Qt.AlignTop | Qt.AlignLeft)
        v_sub.addWidget(self.subject_input)
        
        v_space = QVBoxLayout()

        v_people.addLayout(v_rec)
        v_people.addLayout(v_ccs)
        v_people.addLayout(v_sub)
        v_people.setSpacing(10)
        
        # 2. 邮件正文布局
        v_email_text = QVBoxLayout()
        v_email_text.addWidget(self.email_title )
        v_email_text.addWidget(self.emial_text_input)
        v_email_text.addWidget(self.send_btn, 0, Qt.AlignTop | Qt.AlignRight)
        
        vs.addStretch(3)
        vs.addLayout(v_people)
        vs.addStretch(1)
        vs.addLayout(v_email_text)
        vs.addStretch(3)
        
        self.tab3.setLayout(vs)
        
    def EmailSender(self):
        if len(self.htmls_attachments) == 0:
            self.email_info.append("No report is ready to be sent!")
            return
        receivers = self.receivers_input.text().strip().split(";")
        CCs = self.CCs_input.text()
        subject = self.subject_input.text().strip()
        content = self.emial_text_input.toPlainText().split("\n")
        inputs = [
            receivers,
            CCs,
            self.htmls_attachments,
            subject,
            content
        ]
        self.emial_text_input.setText("")
        self.t3_info("Begin to draft your email ...")
        
        self.thread_email_sender = Thread_Email(inputs)
        self.thread_email_sender._signal.connect(self.t3_info)
        self.thread_email_sender.start()
        time.sleep(2)
        
    def ReconnectDataBase(self):
        global mysql_connect, mysql_cursor, rs_connect, rs_cursor
        try:
            mysql_connect = pymysql.connect(host='dev-dsk-wangting-2a-25bc2431.us-west-2.amazon.com', 
                                      user='avs_user', 
                                      password='avs_pro', 
                                      database='AVS_proactive',
                                      charset='utf8') #服务器名,账户,密码,数据库名
            mysql_cursor = mysql_connect.cursor()

            # 连接 RedShift 数据库
            rs_connect = psycopg2.connect(database="rsbidw",
                                    user="avs_user",
                                    password="A21sP22oR$",
                                    host="rsbi-analytics.clszsz7jap6y.us-east-1.redshift.amazonaws.com",
                                    port="8192")

            rs_cursor = rs_connect.cursor()
        except Exception as e:
            self.print_bug(e)
            self.print_bug('Network Connection Failed')
            return
        self.print_info("数据库连接已刷新！")
        
    def ScoutWeeks(self):
        chosen_year = self.year_path.text()
        try:
            Year = int(chosen_year)
        except:
            self.print_bug("Wrong year!")
            return
        wanted_task = self.wanted_task_path.text().replace(" ","").upper()
        if len(wanted_task) != 5:
            self.print_bug("Unexpected task name!")
            return
        inputs = [wanted_task, chosen_year]
        self.thread_scout = Thread_Scout(inputs)
        self.thread_scout._signal.connect(self.print_info)
        self.thread_scout.start()
        time.sleep(2)
        
    def AddInput(self):
        """
        负责相应 Upload 标签页中添加一组数据的按钮
        """
        wanted_task = self.wanted_task_path.text()
        wanted_week = self.wanted_week_path.text()
        wanted_file = self.wanted_file_path.text()
        
        # 1.Task Name    
        wanted_task = wanted_task.replace(" ","").upper()
        if len(wanted_task) != 5:
            self.t1_bug("Unexpected task name!")
            return
        # 2. Week Num
        if len(wanted_week) == 0:
            self.t1_bug("Unexpected weeks!")
            return
        wanted_week = wanted_week.replace(" ","")
        if "A" in wanted_week or "a" in wanted_week:
            wanted_week = "A"
            self.t1_bug("Seemed like you wanted it all!")
        # 3. File Name
        wanted_file = wanted_file.replace(" ","")
        # 4.是否查重
        wanted_check = "不查重"
        if self.wanted_check.isChecked():
            wanted_check = "查重"
        
        # 输入至信息展示框中
        t = f"{wanted_task}-{wanted_week}-{wanted_file}-{wanted_check}"
        self.input_info.append(t)
        # 清空 3 个输入框的内容
        self.wanted_task_path.setText("JP")
        self.wanted_week_path.setText("")
        self.wanted_file_path.setText("")
        
    def Upload_Touched_thread(self):
        self.t1_info("开始上传 Touched Data ...")
        inputs = self.input_info.toPlainText().strip().split("\n")
        if len(inputs) == 0:
            self.t1_bug("Input is null, cannot upload any data.")
            return
        for i in inputs:
            a,b,c,d = i.split("-")
            
            if b != "A":
                b = [int(num) for num in b.split(",")]
            c = c.split(",")
            if d == "查重":
                wanted_check = True
            else:
                wanted_check = False            
            self.Inputs.append([a,b,c,d])
        self.thread_upload_Touched = Thread_Upload_Touched(self.Inputs)
        self.thread_upload_Touched._signal.connect(self.t1_info)
        self.thread_upload_Touched._bug_signal.connect(self.t1_bug)
        self.thread_upload_Touched._p_signal.connect(self.t1_progress)
        self.thread_upload_Touched.start()
        time.sleep(2)
        
    def Upload_CM_thread(self):
        self.t1_info("开始上传 CM Data ...")
        self.thread_upload_CM = Thread_Upload_CM()
        self.thread_upload_CM._signal.connect(self.t1_info)
        self.thread_upload_CM._bug_signal.connect(self.t1_bug)
        self.thread_upload_CM._p_signal.connect(self.t1_progress)
        self.thread_upload_CM.start()
        time.sleep(2)
        
    def Exit(self):
        self.exit_flag = True
        qApp.quit()
        
    def choose_root_dir(self):
        fname = QFileDialog.getExistingDirectory(self, "Select your root folder", self.root_root_path)
        if fname:
            self.root_path.setText(fname)
    
    def open_source_file(self):
        fname = QFileDialog.getOpenFileName(self, "Open Source File ", "/home")
        if fname:
            self.input_path.setText(fname[0])
        else:
            self.info.insertPlainText("源文件选取有误！")
            return

    def open_CM_info_file(self):
        fname = QFileDialog.getOpenFileName(self, "Open CM info File ", "/home")
        if fname:
            self.CM_path.setText(fname[0])
        else:
            self.info.insertPlainText("CM info 文件选取有误！")
            return
        
    def open_params_file(self):
        fname = QFileDialog.getOpenFileName(self, "Open CM info File ", r"\\ant\dept-as\PEK10\DEPT2\RBS\AVS\AVS_Task Related\Proactive Task\[Updating] Proactive work allocation")
        if fname:
            self.params_file_path.setText(fname[0])
        else:
            self.info.insertPlainText("Params 文件选取有误！")
            return
        
    def set_pbar(self, val:int):
        """
        val: int
        """
        self.pbar.setValue(val)
        
    def t1_bug(self, txt):
        print(txt)
        self.bug_info.append(txt)
        
    def t1_info(self, txt):
        print(txt)
        self.progress_info.append(txt)
        
    def t1_progress(self, num):
        elem = "//\/"*25
        self.t1_pbar.setText(elem[:num])
        
    def t2_html(self, msg):
        self.info.appendHtml(msg)
        self.info_cursor.movePosition(QTextCursor.End)
        
    def t2_text(self, msg):
        print(msg)
        self.info.appendPlainText(msg)
        self.info_cursor.movePosition(QTextCursor.End)
        
    def t3_info(self, text):
        print(text)
        self.email_info.append(text)
        
    def run(self):
        self.root = self.root_path.text()
        self.task_list_dir = self.input_path.text()
        self.CM_data_file_dir = self.CM_path.text()
        self.params_root = self.params_file_path.text()
        self.q_wk_from = int(self.week_from_path.text())
        self.q_wk_to = int(self.week_to_path.text())
        self.q_year = int(self.year_path.text())
        the_CM, CM_login = self.the_CM_path.text(), self.CM_login_path.text()
        self.the_CM = the_CM
        self.CM_login = CM_login
        try:
            os.chdir(self.root)
        except:
            self.info.appendHtml(f"""<font color={color_dict['warning']}>文件导出路径选择错误，请重新选择</font>""")
            return
        try:
            CM_info = pd.read_excel(self.CM_data_file_dir, sheet_name="CMs")
            the_CMs = list(CM_info.Name.values)
            CMs_login = list(CM_info.Login.values)
            self.the_CMs = the_CMs
            self.CMs_login = CMs_login
        except:
            self.info.appendHtml(f"""<font color={color_dict['warning']}>CM info 文件选择错误，请重新选择</font>""")
            return
        
        self.weeks.append([self.q_wk_from, self.q_wk_to])
        self.years.append(self.q_year)
        if self.multiple_weeks_input.toPlainText().strip():
            multiple_data = ex.multiple_weeks_input.toPlainText().split('\n')
            for i in multiple_data:
                try:
                    curr_week_from, curr_week_to, curr_year = i.split(' ')
                    self.weeks.append([int(curr_week_from), int(curr_week_to)])
                    self.years.append(int(curr_year))
                    self.group_cnt += 1
                except:
                    self.info_text("Unexpected data type for Overall & By Task Report!")
                    return
                
        # extract titles & subtitles & explanations from Params file
        try:
            self.titles_3 = pd.read_excel(self.params_root, sheet_name='Titles_3')
        except:
            self.info.appendHtml(f"""<font color={color_dict['warning']}>参数文件路径错误，请重新选择</font>""")
            return
        
        try:
            self.titles_10 = pd.read_excel(self.params_root, sheet_name='Titles_10')
            self.explanation_file_3 = pd.read_excel(self.params_root,sheet_name='Intro_3')
            self.explanation_file_10 = pd.read_excel(self.params_root,sheet_name='Intro_10')
            self.not_touched_df = pd.read_excel(self.params_root, sheet_name='Not_Touched_Status')
            self.viz_params = pd.read_excel(self.params_root, sheet_name='Viz_Params', index_col=0)
        except:
            self.info_text("The params file was changed, please check before we move on.")
            return
        
        self.intro_3 = list(self.explanation_file_3.iloc[:,-1].values)
        self.intro_10 = list(self.explanation_file_10.iloc[:,-1].values)
        self.not_wanted_status = list(self.not_touched_df.iloc[:,0].values)
        
        self.tb_names = list(self.titles_3.Metrics)
        self.tb_subtitles = list(self.titles_3.Explanation)
        self.pb_names = list(self.titles_10.Metrics)
        self.pb_subtitles = list(self.titles_10.Explanation)
        self.tk_names = self.pb_names
        self.tk_subtitles = self.pb_subtitles
        self.oa_names = self.pb_names
        self.oa_subtitles = self.pb_subtitles
        
        
        datas  = [
            self.q_wk_from, self.q_wk_to, self.q_year, 
            self.weeks, self.years, 
            self.the_CM, self.CM_login, 
            self.the_CMs, self.CMs_login,
            self.task_list_dir, self.not_wanted_status
        ]
        self.thread_get_data = Thread_get_data(datas)
        self.thread_get_data._signal.connect(self.set_pbar)
        self.thread_get_data.html_signal.connect(self.t2_html)
        self.thread_get_data.text_signal.connect(self.t2_text)
        self.thread_get_data.start()
        
        time.sleep(5)
        qmut_init.lock()
        qmut_single.lock()
        qmut_multiple.lock()
        self.html_dir, self.data_dir, self.connects, self.single_week_data, self.multiple_weeks_data = self.thread_get_data.return_datas                                                                                                                                                                                                                                                                                                                                                                                                    
        
        qmut_init.unlock()
        qmut_single.unlock()
        qmut_multiple.unlock()
        
    def run_single_week_report(self):
        if not self.single_week_data:
            self.t2_html(f"""<font color={color_dict['warning']}>WARNING: 数据尚未读取!</font>""")
            return
        
        ## 预设参数 -- Time Compare
        total_metrics = 5 # 目前只有 5 个 metrics
        ori_params = list(self.viz_params.loc['Time Compare',:].values)
        ignore_params, default_selected, Grid_width, \
            padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, title_padding = ori_params
        total = total_metrics - ignore_params  # 目前只有 5 个 metrics
        ## 获取其余参数      
        viz_params_time_compare = [
            ignore_params, default_selected, Grid_width,
            padding_vert, padding_hori, grid_vert, grid_hori,
            row_sheet_counts, total_metrics, title_padding
        ]
        
        ## 预设参数 -- By Person
        total_metrics = 10 
        ori_params = list(self.viz_params.loc['By Person',:].values)
        ignore_params, default_selected, Grid_width, \
            padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, title_padding = ori_params
        total = total_metrics - ignore_params # 目前只有 10 个 metrics
        ## 获取其余参数
        viz_params_by_person = [
            ignore_params, default_selected, Grid_width,
            padding_vert, padding_hori, grid_vert, grid_hori,
            row_sheet_counts, total_metrics, title_padding
        ]
            
        datas = [
            self.q_wk_from, self.q_wk_to, self.q_year, 
            self.weeks, self.years, 
            self.the_CM, self.CM_login, 
            self.the_CMs, self.CMs_login,
            self.task_list_dir,
            self.data_dir, self.html_dir,
            self.single_week_data, viz_params_time_compare, viz_params_by_person,
            self.tb_names, self.tb_subtitles,
            self.pb_names, self.pb_subtitles,
            self.intro_3, self.intro_10
        ]
        self.thread_single_week = Thread_single_week(datas)
        self.thread_single_week._signal.connect(self.set_pbar)
        self.thread_single_week._name_signal.connect(self.compile_html_dirs) # 用于发送文件名字
        self.thread_single_week.start()
        
        time.sleep(2)
        qmut_single.lock()
        self.time_compare_grid, self.by_person_grid = self.thread_single_week.return_datas
        self.t2_html(f"""<font color={color_dict['success']}>Congrats! By Person & Time Compare 报告制作完毕!</font>""")
        qmut_single.unlock()
            
    def run_multiple_weeks_report(self):
        if not self.multiple_weeks_data:
            self.t2_html(f"""<font color={color_dict['warning']}>WARNING: 数据尚未读取!</font>""")
            return
        
        ## 预设参数 -- By Task
        total_metrics = 10
        ori_params = list(self.viz_params.loc['By Task',:].values)
        ignore_params, default_selected, Grid_width, \
            padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, title_padding = ori_params
        total = total_metrics - ignore_params
        ## 获取其余参数
        viz_params_by_task = [
            ignore_params, default_selected, Grid_width,
            padding_vert, padding_hori, grid_vert, grid_hori,
            row_sheet_counts, total_metrics, title_padding
        ]
        
        ## 预设参数 -- Overall
        total_metrics = 10
        ori_params = list(self.viz_params.loc['Overall',:].values)
        ignore_params, default_selected, Grid_width, \
            padding_vert, padding_hori, grid_vert, grid_hori, row_sheet_counts, title_padding = ori_params
        total = total_metrics - ignore_params
        ## 获取其余参数
        viz_params_overall = [
            ignore_params, default_selected, Grid_width,
            padding_vert, padding_hori, grid_vert, grid_hori,
            row_sheet_counts, total_metrics, title_padding
        ]
        
        
        datas = [
            self.q_wk_from, self.q_wk_to, self.q_year, 
            self.weeks, self.years, 
            self.the_CM, self.CM_login, 
            self.the_CMs, self.CMs_login,
            self.task_list_dir,
            self.data_dir, self.html_dir,
            self.multiple_weeks_data, viz_params_by_task, viz_params_overall,
            self.tk_names, self.tk_subtitles,
            self.oa_names, self.oa_subtitles, self.intro_10
        ]
        self.thread_multiple_weeks = Thread_multiple_weeks(datas)
        self.thread_multiple_weeks._signal.connect(self.set_pbar)
        self.thread_multiple_weeks._name_signal.connect(self.compile_html_dirs) # 用于发送文件名字
        self.thread_multiple_weeks.start()
        
        time.sleep(2)
        qmut_multiple.lock()
        self.by_task_grid, self.overall_grid = self.thread_multiple_weeks.return_datas
        self.t2_html(f"""<font color={color_dict['success']}>Congrats! By Task & Overall 报告制作完毕!</font>""")
        qmut_multiple.unlock()
        
    def all_report_html(self):
        # 获取日期
        this_moment = time.localtime(time.time())
        curr_year = self.curr_year
        curr_mon = this_moment.tm_mon
        curr_day = this_moment.tm_mday
        # 各个 Tab 的标题
        html_tab_titles = [
            'History File Report', "Time Compare", "By Person", 
            "By Task", "Overall",
            f"History File Report Dashboard_{curr_year}_{curr_mon}_{curr_day}.html"
        ]
        # Input 参数
        datas = [
            self.time_compare_grid, self.by_person_grid,
            self.by_task_grid, self.overall_grid,
            html_tab_titles
        ]
        
        self.integrated_html_grid = html_tab_titles[-1]
        self.compile_html_dirs(self.integrated_html_grid)
        self.thread_integrate = Thread_Integrate(datas)
        self.thread_integrate._signal.connect(self.set_pbar)
        self.thread_integrate.start()
        
        time.sleep(2)
        qmut_integrate.lock()
        self.t2_html(f"""<font color={color_dict['success']}>Congrats! History File Report 制作完毕!</font>""")
        qmut_integrate.unlock()
        
    def compile_html_dirs(self, html_dir):
        res = os.path.abspath('.') + '\\' + html_dir
        self.htmls_attachments.append(res)
        
# 程序入口

if __name__ == '__main__':
    if not QApplication.instance():
        app = QApplication(sys.argv)
    else:
        app = QApplication.instance()
    ex = MYGUI()
    ex.show()
    sys.exit(app.exec_())
    
